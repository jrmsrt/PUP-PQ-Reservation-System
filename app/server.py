from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, Response
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date, time, timedelta
import csv
import json
import re
import urllib.error
import urllib.request
from io import StringIO
from flask_wtf.csrf import generate_csrf
from markupsafe import Markup
from collections import Counter, defaultdict

# Blueprints
main_bp = Blueprint('main', __name__)
auth_bp = Blueprint('auth', __name__)
student_bp = Blueprint('student', __name__)
admin_bp = Blueprint('admin', __name__)
ai_bp = Blueprint('ai', __name__)

# Database Setup & Imports
from app.db import (
    get_db_connection,
    import_student_to_masterlist,
    get_student_by_email,
    get_student_masterlist_entry,
    check_otp_rate_limit,
    check_lockout_active,
    generate_and_save_otp,
    verify_otp_code
)

import smtplib
import threading
import os
from email.mime.text import MIMEText
from email.utils import make_msgid, formatdate

def _env_bool(name, default=False):
    value = os.environ.get(name)
    if value is None:
        return default
    return value.strip().lower() in ('1', 'true', 'yes', 'on')

def _open_smtp_connection(mail_server, mail_port):
    if _env_bool('MAIL_USE_SSL', False):
        server = smtplib.SMTP_SSL(mail_server, mail_port, timeout=20)
    else:
        server = smtplib.SMTP(mail_server, mail_port, timeout=20)
        server.ehlo()
        if _env_bool('MAIL_USE_TLS', True):
            server.starttls()
            server.ehlo()
    return server

def _send_brevo_api_email(recipient_email, subject, body):
    api_key = os.environ.get('BREVO_API_KEY')
    if not api_key:
        return None

    mail_sender = os.environ.get('MAIL_DEFAULT_SENDER') or os.environ.get('MAIL_USERNAME')
    sender_name = os.environ.get('MAIL_SENDER_NAME', 'PUP Reservation System')
    if not mail_sender:
        print("BREVO_API_KEY is configured, but MAIL_DEFAULT_SENDER is missing.", flush=True)
        return False

    payload = {
        'sender': {'name': sender_name, 'email': mail_sender},
        'to': [{'email': recipient_email.strip()}],
        'subject': subject,
        'textContent': body
    }
    req = urllib.request.Request(
        'https://api.brevo.com/v3/smtp/email',
        data=json.dumps(payload).encode('utf-8'),
        headers={
            'accept': 'application/json',
            'api-key': api_key,
            'content-type': 'application/json'
        },
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=20) as response:
            print(f"Brevo API email accepted for {recipient_email}: HTTP {response.status}", flush=True)
            return 200 <= response.status < 300
    except urllib.error.HTTPError as e:
        error_body = e.read().decode('utf-8', errors='replace')
        print(f"Brevo API rejected email to {recipient_email}: HTTP {e.code} {error_body}", flush=True)
        return False
    except Exception as e:
        import sys, traceback
        print(f"Brevo API email error for {recipient_email}: {e}", flush=True)
        traceback.print_exc(file=sys.stdout)
        sys.stdout.flush()
        return False

def _send_otp_email_sync(recipient_email, otp):
    recipient_email = recipient_email.strip()
    mail_server = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
    try:
        mail_port = int(os.environ.get('MAIL_PORT', 587))
    except ValueError:
        mail_port = 587
    mail_username = os.environ.get('MAIL_USERNAME')
    mail_password = os.environ.get('MAIL_PASSWORD')
    mail_sender = os.environ.get('MAIL_DEFAULT_SENDER') or mail_username
    if not mail_username or not mail_password or not mail_sender:
        print("Mail credentials are not configured; OTP email was not sent.", flush=True)
        return False
    
    subject = "PUP Reservation System OTP Verification"
    body = f"Hello,\n\nYour One-Time Password (OTP) for accessing the PUP Reservation System is:\n\n{otp}\n\nThis code will expire in 5 minutes.\n\nIf you did not request this code, please ignore this email.\n\nThank you."
    brevo_result = _send_brevo_api_email(recipient_email, subject, body)
    if brevo_result is not None:
        return brevo_result
    
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = mail_sender
    msg['To'] = recipient_email
    msg['Date'] = formatdate(localtime=True)
    msg['Message-ID'] = make_msgid()
    
    try:
        import sys
        server = _open_smtp_connection(mail_server, mail_port)
        server.login(mail_username, mail_password)
        server.sendmail(mail_sender, [recipient_email], msg.as_string())
        server.close()
        print(f"OTP Email sent successfully to {recipient_email}", flush=True)
        return True
    except Exception as e:
        import sys, traceback
        print(f"Error sending OTP email to {recipient_email}: {e}", flush=True)
        traceback.print_exc(file=sys.stdout)
        sys.stdout.flush()
        return False

def send_otp_email_async(recipient_email, otp):
    if os.environ.get('RENDER') or _env_bool('MAIL_SEND_SYNC', True):
        return _send_otp_email_sync(recipient_email, otp)
    thread = threading.Thread(target=_send_otp_email_sync, args=(recipient_email, otp))
    thread.daemon = True
    thread.start()
    return True

def _send_notification_email_sync(recipient_email, subject, body):
    recipient_email = recipient_email.strip()
    brevo_result = _send_brevo_api_email(recipient_email, subject, body)
    if brevo_result is not None:
        return brevo_result

    mail_server = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
    try:
        mail_port = int(os.environ.get('MAIL_PORT', 587))
    except ValueError:
        mail_port = 587
    mail_username = os.environ.get('MAIL_USERNAME')
    mail_password = os.environ.get('MAIL_PASSWORD')
    mail_sender = os.environ.get('MAIL_DEFAULT_SENDER') or mail_username
    if not mail_username or not mail_password or not mail_sender:
        print("Mail credentials are not configured; notification email was not sent.", flush=True)
        return False
    
    msg = MIMEText(body)
    msg['Subject'] = subject
    msg['From'] = mail_sender
    msg['To'] = recipient_email
    msg['Date'] = formatdate(localtime=True)
    msg['Message-ID'] = make_msgid()
    
    try:
        import sys
        server = _open_smtp_connection(mail_server, mail_port)
        server.login(mail_username, mail_password)
        server.sendmail(mail_sender, [recipient_email], msg.as_string())
        server.close()
        print(f"Notification Email sent successfully to {recipient_email}", flush=True)
        return True
    except Exception as e:
        import sys, traceback
        print(f"Error sending notification email to {recipient_email}: {e}", flush=True)
        traceback.print_exc(file=sys.stdout)
        sys.stdout.flush()
        return False

def send_notification_email_async(recipient_email, subject, body):
    if os.environ.get('RENDER') or _env_bool('MAIL_SEND_SYNC', True):
        return _send_notification_email_sync(recipient_email, subject, body)
    thread = threading.Thread(target=_send_notification_email_sync, args=(recipient_email, subject, body))
    thread.daemon = True
    thread.start()
    return True

LOWERCASE_WORDS = {
    'the', 'on', 'in', 'of', 'or', 'to', 'and', 'at'
}

def capitalize_word(word):
    if '-' in word:
        return '-'.join(capitalize_word(p) for p in word.split('-'))
        
    capitalized = ""
    found_alpha = False
    for char in word:
        if char.isalpha() and not found_alpha:
            capitalized += char.upper()
            found_alpha = True
        elif char.isalpha() and found_alpha:
            capitalized += char.lower()
        else:
            capitalized += char
    return capitalized

def format_course_name(name):
    if not name:
        return name
    
    words = name.split()
    if not words:
        return name
        
    formatted_words = []
    num_words = len(words)
    
    for idx, word in enumerate(words):
        clean_word = word.strip(".,()[]{}!?:;\"'").lower()
        is_first_or_last = (idx == 0 or idx == num_words - 1)
        
        after_colon = False
        if idx > 0:
            prev_word = words[idx-1]
            if prev_word.endswith(':'):
                after_colon = True
                
        if clean_word in LOWERCASE_WORDS and not is_first_or_last and not after_colon:
            formatted_words.append(word.lower())
        else:
            formatted_words.append(capitalize_word(word))
            
    return ' '.join(formatted_words)

USERS_DB = []
FACILITIES_DB = []
PROJECTORS_DB = []
RESERVATIONS_DB = []
system_logs = []

def sync_db_to_memory():
    global USERS_DB, FACILITIES_DB, PROJECTORS_DB, RESERVATIONS_DB, system_logs
    USERS_DB.clear()
    FACILITIES_DB.clear()
    PROJECTORS_DB.clear()
    RESERVATIONS_DB.clear()
    system_logs.clear()
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 1. Load Admin
    cursor.execute("SELECT username, password_hash FROM admin")
    for row in cursor.fetchall():
        USERS_DB.append({
            'id': row['username'],  # Admin ID is 'admin' string
            'email': row['username'] + '@pup.edu.ph',
            'password_hash': row['password_hash'],
            'first_name': 'PUP',
            'middle_name': None,
            'last_name': 'Administrator',
            'contact_number': '09170000000',
            'student_number': None,
            'program': None,
            'year_section': None,
            'role': 'admin',
            'is_active': True,
            'password_changed': True,
            'email_verified': True,
            'created_at': datetime.now()
        })
        
    # 2. Load Users
    cursor.execute('''
        SELECT id AS user_id, student_number, pup_email, password_hash, 
               email_verified, password_changed, account_status, role,
               contact_number, program, year_section, created_at,
               first_name, last_name, middle_name
        FROM authorized_users
    ''')
    for row in cursor.fetchall():
        is_active = (row['account_status'] == 'ACTIVE')
        USERS_DB.append({
            'id': row['user_id'],
            'email': row['pup_email'],
            'password_hash': row['password_hash'],
            'first_name': row['first_name'] or '',
            'middle_name': row['middle_name'] or '',
            'last_name': row['last_name'] or '',
            'contact_number': row['contact_number'] or '',
            'student_number': row['student_number'],
            'program': row['program'] or '',
            'year_section': row['year_section'] or '',
            'role': row['role'].lower() if row['role'] else 'student',
            'is_active': is_active,
            'password_changed': bool(row['password_changed']),
            'email_verified': bool(row['email_verified']),
            'created_at': datetime.strptime(row['created_at'], '%Y-%m-%d %H:%M:%S') if isinstance(row['created_at'], str) else row['created_at']
        })
        
    # 3. Load Facilities
    cursor.execute("SELECT id, code, type, status FROM facilities")
    for row in cursor.fetchall():
        FACILITIES_DB.append({
            'id': row['id'],
            'code': row['code'],
            'type': row['type'],
            'status': row['status']
        })
        
    # 4. Load Projectors
    cursor.execute("SELECT id, code, model, status FROM projectors")
    for row in cursor.fetchall():
        PROJECTORS_DB.append({
            'id': row['id'],
            'code': row['code'],
            'model': row['model'],
            'status': row['status']
        })
        
    # 5. Load Reservations
    cursor.execute("SELECT id, student_id, facility_type, facility_id, projector_id, schedule_date, start_time, end_time, course_code, course_name, professor, status, remarks, checkout_time, return_time, released_by, received_by, returned_to, equipment_condition, created_at FROM reservation_requests")
    for row in cursor.fetchall():
        sd_str = row['schedule_date']
        st_str = row['start_time']
        et_str = row['end_time']
        
        sd = datetime.strptime(sd_str, '%Y-%m-%d').date() if isinstance(sd_str, str) else sd_str
        st = datetime.strptime(st_str, '%H:%M:%S').time() if isinstance(st_str, str) else st_str
        et = datetime.strptime(et_str, '%H:%M:%S').time() if isinstance(et_str, str) else et_str
        
        RESERVATIONS_DB.append({
            'id': row['id'],
            'user_id': row['student_id'],
            'facility_type': row['facility_type'],
            'facility_id': row['facility_id'],
            'projector_id': row['projector_id'],
            'schedule_date': sd,
            'start_time': st,
            'end_time': et,
            'course_code': row['course_code'],
            'course_name': row['course_name'],
            'professor': row['professor'],
            'status': row['status'],
            'remarks': row['remarks'],
            'checkout_time': row['checkout_time'],
            'return_time': row['return_time'],
            'released_by': row['released_by'],
            'received_by': row['received_by'],
            'returned_to': row['returned_to'],
            'equipment_condition': row['equipment_condition'],
            'created_at': datetime.strptime(row['created_at'], '%Y-%m-%d %H:%M:%S') if isinstance(row['created_at'], str) else row['created_at']
        })
        
    # 6. Load System Logs
    cursor.execute("SELECT id, admin_username, student_id, action, details, ip_address, created_at FROM system_logs")
    for row in cursor.fetchall():
        user_id = row['admin_username'] if row['admin_username'] else row['student_id']
        ca_str = row['created_at']
        ca = datetime.strptime(ca_str, '%Y-%m-%d %H:%M:%S') if isinstance(ca_str, str) else ca_str
        
        system_logs.append({
            'id': row['id'],
            'user_id': user_id,
            'action': row['action'],
            'details': row['details'],
            'ip_address': row['ip_address'],
            'created_at': ca
        })
        
    conn.close()

sync_db_to_memory()

@main_bp.before_app_request
def before_request_sync():
    # Data is synced after writes and at startup. Avoid reloading every table on
    # every request; that becomes expensive as logs and reservations grow.
    return

@main_bp.after_app_request
def add_cache_control_headers(response):
    path = request.path
    if path == '/' or path.startswith('/student') or path.startswith('/admin') or path in ('/login', '/verify-otp', '/change-password', '/forgot-password', '/password-decision'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '-1'
    return response

# Helper functions to log system events
def log_event(user_id, action, details):
    conn = get_db_connection()
    cursor = conn.cursor()
    # Check if user_id is admin (string username) or student (integer ID)
    if isinstance(user_id, str) or user_id == 'admin':
        cursor.execute('''
            INSERT INTO system_logs (admin_username, action, details, ip_address) 
            VALUES (?, ?, ?, ?)
        ''', ('admin', action, details, request.remote_addr))
    else:
        cursor.execute('''
            INSERT INTO system_logs (student_id, action, details, ip_address) 
            VALUES (?, ?, ?, ?)
        ''', (user_id, action, details, request.remote_addr))
    conn.commit()
    conn.close()

    system_logs.append({
        'id': len(system_logs) + 1,
        'user_id': user_id,
        'action': action,
        'details': details,
        'ip_address': request.remote_addr,
        'created_at': datetime.now()
    })

# Emulated Models
class MockUser:
    def __init__(self, user_dict=None):
        if user_dict:
            self.id = user_dict.get('id')
            self.email = user_dict.get('email')
            self.first_name = user_dict.get('first_name')
            self.middle_name = user_dict.get('middle_name')
            self.last_name = user_dict.get('last_name')
            self.contact_number = user_dict.get('contact_number') or ''
            self.student_number = user_dict.get('student_number') or ''
            self.program = user_dict.get('program') or ''
            self.year_section = user_dict.get('year_section') or ''
            self.role = user_dict.get('role').lower() if user_dict.get('role') else None
            self.is_active = user_dict.get('is_active', True)
            self.password_changed = user_dict.get('password_changed', False)
            self.email_verified = user_dict.get('email_verified', False)
            self.created_at = user_dict.get('created_at')
            self.is_authenticated = True
        else:
            self.id = None
            self.role = 'guest'
            self.is_authenticated = False
            self.first_name = ''
            self.middle_name = ''
            self.last_name = ''
            self.email = ''
            self.student_number = ''
            self.created_at = None
            
    @property
    def full_name(self):
        if self.role == 'admin':
            return f"{self.first_name} {self.last_name}"
        m_name = getattr(self, 'middle_name', '')
        if m_name and m_name.strip():
            m_initial = m_name.strip()[0].upper()
            return f"{self.last_name}, {self.first_name} {m_initial}."
        return f"{self.last_name}, {self.first_name}"

class FacilityObj:
    def __init__(self, d):
        self.id = d.get('id')
        self.code = d.get('code')
        self.type = d.get('type')
        self.status = d.get('status')

class ProjectorObj:
    def __init__(self, d):
        self.id = d.get('id')
        self.code = d.get('code')
        self.model = d.get('model')
        self.status = d.get('status')

class ReservationObj:
    def __init__(self, res_dict):
        self.id = res_dict.get('id')
        self.user_id = res_dict.get('user_id')
        self.facility_type = res_dict.get('facility_type')
        self.facility_id = res_dict.get('facility_id')
        self.projector_id = res_dict.get('projector_id')
        
        sd = res_dict.get('schedule_date')
        if isinstance(sd, str):
            self.schedule_date = datetime.strptime(sd, '%Y-%m-%d').date()
        else:
            self.schedule_date = sd
            
        st = res_dict.get('start_time')
        if isinstance(st, str):
            self.start_time = datetime.strptime(st, '%H:%M').time()
        elif isinstance(st, datetime):
            self.start_time = st.time()
        else:
            self.start_time = st
            
        et = res_dict.get('end_time')
        if isinstance(et, str):
            self.end_time = datetime.strptime(et, '%H:%M').time()
        elif isinstance(et, datetime):
            self.end_time = et.time()
        else:
            self.end_time = et
            
        self.course_code = res_dict.get('course_code')
        self.course_name = res_dict.get('course_name')
        self.professor = res_dict.get('professor')
        self.status = res_dict.get('status')
        self.remarks = res_dict.get('remarks')
        self.checkout_time = res_dict.get('checkout_time')
        self.return_time = res_dict.get('return_time')
        self.released_by = res_dict.get('released_by')
        self.received_by = res_dict.get('received_by')
        self.returned_to = res_dict.get('returned_to')
        self.equipment_condition = res_dict.get('equipment_condition')
        self.created_at = res_dict.get('created_at')

    @property
    def user(self):
        for u in USERS_DB:
            if u['id'] == self.user_id:
                return MockUser(u)
        return None

    @property
    def facility(self):
        if self.facility_id:
            for f in FACILITIES_DB:
                if f['id'] == self.facility_id:
                    return FacilityObj(f)
        return None

    @property
    def projector(self):
        if self.projector_id:
            for p in PROJECTORS_DB:
                if p['id'] == self.projector_id:
                    return ProjectorObj(p)
        return None

class SystemLogObj:
    def __init__(self, d):
        self.id = d.get('id')
        self.user_id = d.get('user_id')
        self.action = d.get('action')
        self.details = d.get('details')
        self.ip_address = d.get('ip_address')
        self.created_at = d.get('created_at')
        
    @property
    def user(self):
        if self.user_id:
            for u in USERS_DB:
                if u['id'] == self.user_id:
                    return MockUser(u)
        return None

# Emulated Form Fields for WTF emulating
class MockFormField:
    def __init__(self, data=None, errors=None):
        self.data = data or ''
        self.errors = errors or []
    def __call__(self, **kwargs):
        attrs = ' '.join(f'{k}="{v}"' for k, v in kwargs.items())
        return Markup(f'<input type="file" name="file" {attrs}>')

class MockForm:
    def __init__(self, **kwargs):
        self.csrf_token = Markup(f'<input type="hidden" name="csrf_token" value="{generate_csrf()}">')
        for k, v in kwargs.items():
            setattr(self, k, MockFormField(data=v))
            
    def __getattr__(self, name):
        return MockFormField()

def get_current_user():
    user_id = session.get('user_id')
    if not user_id:
        return MockUser(None)
        
    if str(user_id) == 'admin':
        # Retrieve admin from USERS_DB (seeded at startup)
        for u in USERS_DB:
            if u['id'] == 'admin':
                return MockUser(u)
        return MockUser(None)
        
    # Query database directly for student
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id AS user_id, student_number, pup_email, password_hash, 
                   email_verified, password_changed, account_status, role,
                   contact_number, program, year_section, created_at,
                   first_name, last_name, middle_name
            FROM authorized_users
            WHERE id = ?
        ''', (user_id,))
        row = cursor.fetchone()
        conn.close()
        
        if row:
            is_active = (row['account_status'] == 'ACTIVE')
            user_dict = {
                'id': row['user_id'],
                'email': row['pup_email'],
                'password_hash': row['password_hash'],
                'first_name': row['first_name'] or '',
                'middle_name': row['middle_name'] or '',
                'last_name': row['last_name'] or '',
                'contact_number': row['contact_number'] or '',
                'student_number': row['student_number'],
                'program': row['program'] or '',
                'year_section': row['year_section'] or '',
                'role': row['role'].lower() if row['role'] else 'student',
                'is_active': is_active,
                'password_changed': bool(row['password_changed']),
                'email_verified': bool(row['email_verified']),
                'created_at': row['created_at']
            }
            return MockUser(user_dict)
    except Exception as e:
        import sys
        print(f"Error in get_current_user: {e}", file=sys.stderr, flush=True)
        
    return MockUser(None)

def get_mock_form():
    u = get_current_user()
    if u.is_authenticated:
        return MockForm(
            contact_number=u.contact_number,
            program=u.program,
            year_section=u.year_section
        )
    return MockForm()

def _reservation_lock_key(facility_id, projector_id, schedule_date):
    resource_kind = 'facility' if facility_id else 'projector'
    resource_id = facility_id or projector_id
    return f"reservation:{resource_kind}:{resource_id}:{schedule_date}"

def _acquire_reservation_lock(cursor, facility_id, projector_id, schedule_date):
    cursor.execute("SELECT GET_LOCK(?, 5)", (_reservation_lock_key(facility_id, projector_id, schedule_date),))
    row = cursor.fetchone()
    return bool(row and row[0] == 1)

def _release_reservation_lock(cursor, facility_id, projector_id, schedule_date):
    cursor.execute("SELECT RELEASE_LOCK(?)", (_reservation_lock_key(facility_id, projector_id, schedule_date),))

def has_reservation_conflict(cursor, schedule_date, start_time, end_time, facility_id=None, projector_id=None,
                             statuses=('APPROVED', 'PENDING APPROVAL', 'IN USE'), exclude_id=None):
    if not facility_id and not projector_id:
        return False

    status_placeholders = ', '.join(['?'] * len(statuses))
    resource_clause = "facility_id = ?" if facility_id else "projector_id = ?"
    params = list(statuses) + [
        schedule_date.strftime('%Y-%m-%d') if hasattr(schedule_date, 'strftime') else schedule_date,
        end_time.strftime('%H:%M:%S') if hasattr(end_time, 'strftime') else end_time,
        start_time.strftime('%H:%M:%S') if hasattr(start_time, 'strftime') else start_time,
        facility_id or projector_id
    ]

    exclude_clause = ""
    if exclude_id is not None:
        exclude_clause = "AND id != ?"
        params.append(exclude_id)

    cursor.execute(f'''
        SELECT id FROM reservation_requests
        WHERE status IN ({status_placeholders})
          AND schedule_date = ?
          AND start_time < ?
          AND end_time > ?
          AND {resource_clause}
          {exclude_clause}
        LIMIT 1
    ''', tuple(params))
    return cursor.fetchone() is not None

def _format_reservation_schedule_for_email(res_row):
    try:
        s_date = res_row['schedule_date']
        if isinstance(s_date, str):
            s_date_str = datetime.strptime(s_date, '%Y-%m-%d').strftime('%B %d, %Y')
        else:
            s_date_str = s_date.strftime('%B %d, %Y')

        s_time = res_row['start_time']
        if isinstance(s_time, str):
            s_time_obj = datetime.strptime(s_time, '%H:%M:%S' if s_time.count(':') == 2 else '%H:%M')
            s_time_str = s_time_obj.strftime('%I:%M %p')
        else:
            s_time_str = s_time.strftime('%I:%M %p')

        e_time = res_row['end_time']
        if isinstance(e_time, str):
            e_time_obj = datetime.strptime(e_time, '%H:%M:%S' if e_time.count(':') == 2 else '%H:%M')
            e_time_str = e_time_obj.strftime('%I:%M %p')
        else:
            e_time_str = e_time.strftime('%I:%M %p')
    except Exception as e:
        print("Error parsing datetime for email:", e)
        s_date_str = str(res_row['schedule_date'])
        s_time_str = str(res_row['start_time'])
        e_time_str = str(res_row['end_time'])

    return s_date_str, s_time_str, e_time_str

def _get_reservation_resource_code(res_row):
    code = 'N/A'
    conn = get_db_connection()
    cursor = conn.cursor()
    if res_row['facility_id']:
        cursor.execute("SELECT code FROM facilities WHERE id = ?", (res_row['facility_id'],))
        row = cursor.fetchone()
        if row:
            code = row['code']
    elif res_row['projector_id']:
        cursor.execute("SELECT code FROM projectors WHERE id = ?", (res_row['projector_id'],))
        row = cursor.fetchone()
        if row:
            code = row['code']
    conn.close()
    return code

def _get_reserved_projector_ids():
    return {
        r['projector_id']
        for r in RESERVATIONS_DB
        if r.get('projector_id') and r.get('status') in ('APPROVED', 'AWAITING RETURN', 'OVERDUE')
    }

PROJECTOR_HARD_BLOCK_STATUSES = {'Under Maintenance', 'Maintenance', 'Unavailable', 'Checked Out', 'Overdue', 'Archived'}
ACTIVE_PROJECTOR_RESERVATION_STATUSES = ('APPROVED', 'AWAITING RETURN', 'OVERDUE')
FACILITY_HARD_BLOCK_STATUSES = {'Maintenance', 'Unavailable', 'Archived'}
ACTIVE_FACILITY_RESERVATION_STATUSES = ('APPROVED', 'IN USE')

def _reservation_end_datetime(res):
    schedule_date = res['schedule_date']
    end_time = res['end_time']
    if isinstance(schedule_date, str):
        schedule_date = datetime.strptime(schedule_date, '%Y-%m-%d').date()
    if isinstance(end_time, str):
        end_time = datetime.strptime(end_time, '%H:%M:%S' if end_time.count(':') == 2 else '%H:%M').time()
    return datetime.combine(schedule_date, end_time)

def _active_projector_reservation_exists(cursor, projector_id, exclude_id=None):
    params = [projector_id]
    exclude_clause = ""
    if exclude_id is not None:
        exclude_clause = "AND id != ?"
        params.append(exclude_id)
    cursor.execute(f'''
        SELECT id
        FROM reservation_requests
        WHERE projector_id = ?
          AND status IN ('APPROVED', 'AWAITING RETURN', 'OVERDUE')
          AND return_time IS NULL
          {exclude_clause}
        LIMIT 1
    ''', tuple(params))
    return cursor.fetchone() is not None

def _release_projector_if_free(cursor, projector_id, exclude_reservation_id=None):
    if not projector_id:
        return
    if not _active_projector_reservation_exists(cursor, projector_id, exclude_reservation_id):
        cursor.execute("UPDATE projectors SET status = 'Available' WHERE id = ?", (projector_id,))

def refresh_projector_custody_statuses():
    now = datetime.now()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, projector_id, schedule_date, end_time, status, checkout_time, return_time
        FROM reservation_requests
        WHERE projector_id IS NOT NULL
          AND status IN ('APPROVED', 'AWAITING RETURN', 'OVERDUE')
          AND return_time IS NULL
    ''')
    changed = False
    for row in cursor.fetchall():
        if now <= _reservation_end_datetime(row):
            continue
        if row['checkout_time']:
            if row['status'] != 'OVERDUE':
                cursor.execute("UPDATE reservation_requests SET status = 'OVERDUE' WHERE id = ?", (row['id'],))
                cursor.execute("UPDATE projectors SET status = 'Checked Out' WHERE id = ?", (row['projector_id'],))
                changed = True
        elif row['status'] == 'APPROVED':
            cursor.execute("UPDATE reservation_requests SET status = 'AWAITING RETURN' WHERE id = ?", (row['id'],))
            cursor.execute("UPDATE projectors SET status = 'Reserved' WHERE id = ?", (row['projector_id'],))
            changed = True
    if changed:
        conn.commit()
    conn.close()
    if changed:
        sync_db_to_memory()

def _active_facility_reservation_exists(cursor, facility_id, exclude_id=None):
    params = [facility_id]
    exclude_clause = ""
    if exclude_id is not None:
        exclude_clause = "AND id != ?"
        params.append(exclude_id)
    cursor.execute(f'''
        SELECT id
        FROM reservation_requests
        WHERE facility_id = ?
          AND status IN ('APPROVED', 'IN USE')
          {exclude_clause}
        LIMIT 1
    ''', tuple(params))
    return cursor.fetchone() is not None

def _sync_facility_status_from_reservations(cursor, facility_id):
    cursor.execute('''
        SELECT status
        FROM reservation_requests
        WHERE facility_id = ?
          AND status IN ('APPROVED', 'IN USE')
    ''', (facility_id,))
    statuses = {row['status'] for row in cursor.fetchall()}
    if 'IN USE' in statuses:
        cursor.execute("UPDATE facilities SET status = 'Occupied' WHERE id = ?", (facility_id,))
    elif 'APPROVED' in statuses:
        cursor.execute("UPDATE facilities SET status = 'Reserved' WHERE id = ?", (facility_id,))
    else:
        cursor.execute("UPDATE facilities SET status = 'Available' WHERE id = ?", (facility_id,))

def _release_facility_if_free(cursor, facility_id, exclude_reservation_id=None):
    if not facility_id:
        return
    if not _active_facility_reservation_exists(cursor, facility_id, exclude_reservation_id):
        cursor.execute("UPDATE facilities SET status = 'Available' WHERE id = ?", (facility_id,))

def refresh_facility_usage_statuses():
    now = datetime.now()
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT r.id, r.facility_id, r.schedule_date, r.start_time, r.end_time, r.status,
               f.status AS facility_status
        FROM reservation_requests r
        JOIN facilities f ON r.facility_id = f.id
        WHERE r.facility_id IS NOT NULL
          AND r.status IN ('APPROVED', 'IN USE')
    ''')
    changed = False
    affected_facility_ids = set()
    for row in cursor.fetchall():
        start_dt = datetime.combine(
            row['schedule_date'] if not isinstance(row['schedule_date'], str) else datetime.strptime(row['schedule_date'], '%Y-%m-%d').date(),
            row['start_time'] if not isinstance(row['start_time'], str) else datetime.strptime(row['start_time'], '%H:%M:%S' if row['start_time'].count(':') == 2 else '%H:%M').time()
        )
        end_dt = _reservation_end_datetime(row)

        if now >= end_dt:
            if row['status'] != 'COMPLETED':
                cursor.execute("UPDATE reservation_requests SET status = 'COMPLETED' WHERE id = ?", (row['id'],))
                affected_facility_ids.add(row['facility_id'])
                changed = True
        elif now >= start_dt:
            if row['status'] != 'IN USE':
                cursor.execute("UPDATE reservation_requests SET status = 'IN USE' WHERE id = ?", (row['id'],))
                affected_facility_ids.add(row['facility_id'])
                changed = True
        elif row['status'] == 'APPROVED':
            if row['facility_status'] != 'Reserved':
                affected_facility_ids.add(row['facility_id'])
                changed = True
    for facility_id in affected_facility_ids:
        _sync_facility_status_from_reservations(cursor, facility_id)
    if changed:
        conn.commit()
    conn.close()
    if changed:
        sync_db_to_memory()

DEFAULT_AI_SETTINGS = {
    'buffer_minutes': 15,
    'auto_suggest': True,
    'peak_warning': True,
    'email_alerts': True
}

def load_ai_settings():
    settings = dict(DEFAULT_AI_SETTINGS)
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT setting_key, setting_value FROM ai_settings")
    for row in cursor.fetchall():
        key = row['setting_key']
        value = row['setting_value']
        if key == 'buffer_minutes':
            try:
                settings[key] = max(0, min(60, int(value)))
            except ValueError:
                settings[key] = DEFAULT_AI_SETTINGS[key]
        elif key in ('auto_suggest', 'peak_warning', 'email_alerts'):
            settings[key] = value == '1'
    conn.close()
    return settings

def save_ai_settings(buffer_minutes, auto_suggest, peak_warning, email_alerts):
    values = {
        'buffer_minutes': str(max(0, min(60, int(buffer_minutes)))),
        'auto_suggest': '1' if auto_suggest else '0',
        'peak_warning': '1' if peak_warning else '0',
        'email_alerts': '1' if email_alerts else '0'
    }
    conn = get_db_connection()
    cursor = conn.cursor()
    for key, value in values.items():
        cursor.execute('''
            INSERT INTO ai_settings (setting_key, setting_value)
            VALUES (?, ?)
            ON DUPLICATE KEY UPDATE setting_value = VALUES(setting_value)
        ''', (key, value))
    conn.commit()
    conn.close()


# ==========================================
# Core Main Landing Route
# ==========================================
@main_bp.route('/')
def index():
    u = get_current_user()
    if u.is_authenticated:
        if u.role == 'admin':
            return redirect(url_for('admin.dashboard'))
        else:
            return redirect(url_for('student.dashboard'))
    return render_template('landing.html')


# ==========================================
# Authentication Routes
# ==========================================
@auth_bp.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        session.pop('verified_otp_user_id', None)
        session.pop('reset_password_allowed', None)

    u = get_current_user()
    if u.is_authenticated:
        if u.role == 'admin':
            return redirect(url_for('admin.dashboard'))
        else:
            return redirect(url_for('student.dashboard'))
            
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '').strip()
        role = request.form.get('role', 'student')
        
        if role == 'admin':
            # Admin Login logic
            user = None
            for u in USERS_DB:
                if u['role'] == 'admin' and (u['email'] == email or u['email'].split('@')[0] == email):
                    user = u
                    break
            if user and check_password_hash(user['password_hash'], password):
                session['user_id'] = user['id']
                log_event(user['id'], "LOGIN", "Administrator logged in successfully.")
                flash("Logged in successfully!", "success")
                return redirect(url_for('admin.dashboard'))
            else:
                flash("Invalid credentials.", "danger")
                return render_template('auth/login.html', form=MockForm(email=email))
                
        # Student Login logic
        # 1. Verify Student Eligibility
        masterlist_entry = get_student_masterlist_entry(email)
        if not masterlist_entry:
            flash("Access Denied.\n\nYour PUP Webmail is not included in the authorized student masterlist.\nOnly authorized PUP students may access this system.", "danger")
            return render_template('auth/login.html', form=MockForm(email=email))
            
        # 2. Get user account and check credential
        user_record = get_student_by_email(email)
        if not user_record:
            # If in masterlist but no user account
            flash("Student record found in masterlist, but no login account exists. Contact administrator.", "danger")
            return render_template('auth/login.html', form=MockForm(email=email))
            
        if not check_password_hash(user_record['password_hash'], password):
            flash("Invalid PUP Webmail or password.", "danger")
            return render_template('auth/login.html', form=MockForm(email=email))
            
        if user_record['account_status'] != 'ACTIVE':
            flash("Your account is deactivated. Please contact the administrator.", "danger")
            return render_template('auth/login.html', form=MockForm(email=email))
            
        # Check lockout
        is_locked, lockout_time = check_lockout_active(user_record['user_id'])
        if is_locked:
            flash("Your account is temporarily locked due to too many failed OTP attempts. Please try again after 15 minutes.", "danger")
            return render_template('auth/login.html', form=MockForm(email=email))
            
        # 3. Determine Authentication Path
        if not user_record['password_changed']:
            # Scenario A: Student is still using default password -> Require OTP
            # Check OTP request rate limits
            if check_otp_rate_limit(user_record['user_id']):
                flash("Maximum OTP requests per hour exceeded. Please try again later.", "danger")
                return render_template('auth/login.html', form=MockForm(email=email))
                
            otp = generate_and_save_otp(user_record['user_id'])
            
            if not send_otp_email_async(email, otp):
                flash("The email provider did not accept the OTP message. Please contact the administrator.", "danger")
                return render_template('auth/login.html', form=MockForm(email=email))
            
            log_event(user_record['user_id'], "OTP REQUEST", "Generated secure OTP and sent email verification code.")
            flash("An OTP verification code has been sent to your PUP Webmail.", "info")
            
            session['pending_otp_user_id'] = user_record['user_id']
            return redirect(url_for('auth.verify_otp'))
        else:
            # Scenario B: Student already changed their password -> Skip OTP
            session['user_id'] = user_record['user_id']
            log_event(user_record['user_id'], "LOGIN", "Student logged in successfully (skipped OTP).")
            flash("Logged in successfully!", "success")
            return redirect(url_for('student.dashboard'))
            
    return render_template('auth/login.html', form=MockForm())

@auth_bp.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    u = get_current_user()
    if u.is_authenticated:
        if u.role == 'admin':
            return redirect(url_for('admin.dashboard'))
        else:
            return redirect(url_for('student.dashboard'))
            
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        
        user_record = get_student_by_email(email)
        if not user_record:
            return render_template('auth/forgot_password.html', form=MockForm(), error_msg="Student webmail not found in our database.")
            
        if user_record['account_status'] != 'ACTIVE':
            return render_template('auth/forgot_password.html', form=MockForm(), error_msg="Your account is deactivated. Please contact the administrator.")
            
        is_locked, lockout_time = check_lockout_active(user_record['user_id'])
        if is_locked:
            flash("Your account is temporarily locked due to too many failed OTP attempts. Please try again after 15 minutes.", "danger")
            return redirect(url_for('auth.login'))
            
        if not user_record['email_verified'] or not user_record['password_changed']:
            flash("Your account has not been activated yet. Please use the default password (PUPrs@1904) to log in.", "info")
            return redirect(url_for('auth.login'))
            
        if check_otp_rate_limit(user_record['user_id']):
            return render_template('auth/forgot_password.html', form=MockForm(), error_msg="Maximum OTP requests per hour exceeded. Please try again later.")
            
        otp = generate_and_save_otp(user_record['user_id'])
        if not send_otp_email_async(email, otp):
            return render_template('auth/forgot_password.html', form=MockForm(), error_msg="The email provider did not accept the password reset OTP. Please contact the administrator.")
        
        log_event(user_record['user_id'], "FORGOT PASSWORD OTP", "Requested password reset OTP.")
        flash("A password reset OTP verification code has been sent to your PUP Webmail.", "info")
        
        session['pending_otp_user_id'] = user_record['user_id']
        session['otp_purpose'] = 'forgot_password'
        return redirect(url_for('auth.verify_otp'))
        
    return render_template('auth/forgot_password.html', form=MockForm())

@auth_bp.route('/verify-otp', methods=['GET', 'POST'])
def verify_otp():
    user_id = session.get('pending_otp_user_id')
    if not user_id:
        flash("No login attempt in progress. Please log in first.", "warning")
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        otp_code = request.form.get('otp_code', '').strip()
        
        # Check lockout status
        is_locked, lockout_time = check_lockout_active(user_id)
        if is_locked:
            flash("Your account is temporarily locked due to too many failed OTP attempts. Please try again after 15 minutes.", "danger")
            return redirect(url_for('auth.login'))
            
        success, msg = verify_otp_code(user_id, otp_code)
        sync_db_to_memory()
        
        if success:
            log_event(user_id, "OTP VERIFY SUCCESS", "Successfully verified OTP code.")
            session['verified_otp_user_id'] = user_id
            session.pop('pending_otp_user_id', None)
            
            # Check if this OTP verification is for Forgot Password flow
            if session.get('otp_purpose') == 'forgot_password':
                session['reset_password_allowed'] = True
                session.pop('otp_purpose', None)
                return redirect(url_for('auth.change_password'))
                
            return redirect(url_for('auth.password_decision'))
        else:
            log_event(user_id, "OTP VERIFY FAILED", f"Failed OTP verification: {msg}")
            # If the failure caused a lockout, redirect to login
            is_locked, lockout_time = check_lockout_active(user_id)
            if is_locked:
                session.pop('pending_otp_user_id', None)
                flash("Too many failed attempts. Your account has been temporarily locked for 15 minutes.", "danger")
                return redirect(url_for('auth.login'))
                
            return render_template('auth/verify_otp.html', form=MockForm(), error_msg=msg)
            
    return render_template('auth/verify_otp.html', form=MockForm())

@auth_bp.route('/resend-otp', methods=['POST'])
def resend_otp():
    user_id = session.get('pending_otp_user_id')
    if not user_id:
        flash("No login attempt in progress.", "warning")
        return redirect(url_for('auth.login'))
        
    is_locked, lockout_time = check_lockout_active(user_id)
    if is_locked:
        flash("Your account is locked. Please wait.", "danger")
        return redirect(url_for('auth.login'))
        
    if check_otp_rate_limit(user_id):
        flash("Maximum OTP requests per hour exceeded. Please try again later.", "danger")
        return render_template('auth/verify_otp.html', form=MockForm(), error_msg="OTP request rate limit exceeded.")
        
    otp = generate_and_save_otp(user_id)
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT pup_email FROM authorized_users WHERE id = ?", (user_id,))
    email = cursor.fetchone()['pup_email']
    conn.close()
    
    if not send_otp_email_async(email, otp):
        flash("The email provider did not accept the new OTP message. Please contact the administrator.", "danger")
        return redirect(url_for('auth.verify_otp'))
    
    log_event(user_id, "OTP RESEND", "Resent secure OTP and sent email verification code.")
    flash("A new OTP code has been sent.", "info")
    return redirect(url_for('auth.verify_otp'))

@auth_bp.route('/password-decision', methods=['GET', 'POST'])
def password_decision():
    user_id = session.get('verified_otp_user_id')
    if not user_id:
        flash("Unauthorized access attempt.", "warning")
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        session['user_id'] = user_id
        session.pop('verified_otp_user_id', None)
        log_event(user_id, "PASSWORD DECISION LATER", "User decided to change password later.")
        flash("Logged in successfully using default password. For better security, please change your password from Account Settings.", "warning")
        return redirect(url_for('student.dashboard'))
        
    return render_template('auth/password_decision.html', form=MockForm())

@auth_bp.route('/change-password', methods=['GET', 'POST'])
def change_password():
    user_id = session.get('verified_otp_user_id')
    if not user_id:
        user_id = session.get('user_id')
        
    if not user_id:
        flash("Unauthorized access attempt.", "warning")
        return redirect(url_for('auth.login'))
        
    is_reset = session.get('reset_password_allowed')
        
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute("SELECT password_hash FROM authorized_users WHERE id = ?", (user_id,))
        row = cursor.fetchone()
        if not row:
            conn.close()
            return render_template('auth/change_password.html', form=MockForm(), error_msg="Unable to verify your account password.", hide_current=is_reset)

        if not is_reset and not check_password_hash(row['password_hash'], current_password):
            conn.close()
            return render_template('auth/change_password.html', form=MockForm(), current_password_error="Current password is incorrect.", hide_current=is_reset)
            
        val = new_password
        has_length = len(val) >= 8
        has_upper = any(c.isupper() for c in val)
        has_lower = any(c.islower() for c in val)
        has_number = any(c.isdigit() for c in val)
        has_special = any(not c.isalnum() for c in val)
        
        if not (has_length and has_upper and has_lower and has_number and has_special):
            conn.close()
            return render_template('auth/change_password.html', form=MockForm(), error_msg="New password does not meet requirements.", hide_current=is_reset)

        if check_password_hash(row['password_hash'], new_password):
            conn.close()
            return render_template('auth/change_password.html', form=MockForm(), error_msg="New password cannot be the same as your current password.", hide_current=is_reset)
            
        if new_password != confirm_password:
            conn.close()
            return render_template('auth/change_password.html', form=MockForm(), error_msg="Passwords should match.", hide_current=is_reset)
            
        new_hash = generate_password_hash(new_password)
        cursor.execute("UPDATE authorized_users SET password_hash = ?, password_changed = 1, email_verified = 1 WHERE id = ?", (new_hash, user_id))
        conn.commit()
        conn.close()
        
        sync_db_to_memory()
        
        log_event(user_id, "CHANGE PASSWORD SUCCESS", "Successfully changed password.")
        
        if is_reset:
            session.pop('reset_password_allowed', None)
            session.pop('verified_otp_user_id', None)
            flash("Password reset successfully. You can now log in with your new password.", "success")
            return redirect(url_for('auth.login'))
            
        if session.get('verified_otp_user_id'):
            session['user_id'] = user_id
            session.pop('verified_otp_user_id', None)
            
        flash("Password updated successfully.", "success")
        return redirect(url_for('student.dashboard'))
        
    return render_template('auth/change_password.html', form=MockForm(), hide_current=is_reset)

@auth_bp.route('/logout', methods=['POST'])
def logout():
    u = get_current_user()
    if u.is_authenticated:
        log_event(u.id, "LOGOUT", "User logged out successfully.")
    session.clear()
    flash("Logged out successfully.", "success")
    return redirect(url_for('main.index'))


# ==========================================
# Student Routes
# ==========================================
@student_bp.route('/student/dashboard')
def dashboard():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'student':
        return redirect(url_for('auth.login'))

    refresh_facility_usage_statuses()

    student_res = [r for r in RESERVATIONS_DB if r['user_id'] == u.id]
    total = len(student_res)
    pending = len([r for r in student_res if r['status'] == 'PENDING APPROVAL'])
    approved = len([r for r in student_res if r['status'] == 'APPROVED'])
    rejected = len([r for r in student_res if r['status'] == 'REJECTED'])
    
    recent_dicts = sorted(student_res, key=lambda x: x['created_at'], reverse=True)[:5]
    recent = [ReservationObj(r) for r in recent_dicts]
    
    return render_template('student/dashboard.html', total=total, pending=pending, approved=approved, rejected=rejected, recent=recent)

@student_bp.route('/student/check-facilities')
def check_facilities():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'student':
        return jsonify({'error': 'Unauthorized'}), 401

    refresh_facility_usage_statuses()

    facility_type = request.args.get('type')
    date_str = request.args.get('date')
    start_str = request.args.get('start')
    end_str = request.args.get('end')
    
    if not all([facility_type, date_str, start_str, end_str]):
        return jsonify({'error': 'Missing parameters'}), 400
        
    req_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    req_start = datetime.strptime(start_str, '%H:%M').time()
    req_end = datetime.strptime(end_str, '%H:%M').time()
    
    category_rooms = [f for f in FACILITIES_DB if f['type'] == facility_type and f['status'] != 'Archived']
    type_rooms = [f for f in category_rooms if f['status'] not in FACILITY_HARD_BLOCK_STATUSES]
    available = []
    unavailable = []
    
    for room in type_rooms:
        # Check overlaps
        has_overlap = False
        for res in RESERVATIONS_DB:
            if res['status'] in ['APPROVED', 'PENDING APPROVAL', 'IN USE'] and res['facility_id'] == room['id']:
                res_date = res['schedule_date']
                res_start = res['start_time']
                res_end = res['end_time']
                if res_date == req_date:
                    if req_start < res_end and req_end > res_start:
                        has_overlap = True
                        break
        if has_overlap:
            unavailable.append(room)
        else:
            available.append(room)
            
    blocked_rooms = [
        dict(room, reason=room['status'])
        for room in category_rooms
        if room['status'] in FACILITY_HARD_BLOCK_STATUSES
    ]
    unavailable.extend(blocked_rooms)

    return jsonify({
        'available': available,
        'unavailable': unavailable,
        'total': len(category_rooms),
        'facility_type': facility_type
    })

@student_bp.route('/student/check-projectors')
def check_projectors():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'student':
        return jsonify({'error': 'Unauthorized'}), 401

    date_str = request.args.get('date')
    start_str = request.args.get('start')
    end_str = request.args.get('end')
    
    if not all([date_str, start_str, end_str]):
        return jsonify({'error': 'Missing parameters'}), 400
        
    req_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    req_start = datetime.strptime(start_str, '%H:%M').time()
    req_end = datetime.strptime(end_str, '%H:%M').time()
    
    out_projectors = []
    for proj in PROJECTORS_DB:
        if proj['status'] in PROJECTOR_HARD_BLOCK_STATUSES:
            out_projectors.append({'id': proj['id'], 'code': proj['code'], 'status': proj['status']})
            continue
            
        # Check overlaps
        has_overlap = False
        for res in RESERVATIONS_DB:
            if res['status'] in ['APPROVED', 'PENDING APPROVAL'] and res['projector_id'] == proj['id']:
                res_date = res['schedule_date']
                res_start = res['start_time']
                res_end = res['end_time']
                if res_date == req_date:
                    if req_start < res_end and req_end > res_start:
                        has_overlap = True
                        break
        status = 'Reserved' if has_overlap else 'Available'
        out_projectors.append({'id': proj['id'], 'code': proj['code'], 'status': status})
        
    return jsonify({'projectors': out_projectors})

@student_bp.route('/student/reserve', methods=['GET', 'POST'])
def reserve():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'student':
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        facility_type = request.form.get('facility_type')
        date_str = request.form.get('schedule_date')
        start_str = request.form.get('start_time')
        end_str = request.form.get('end_time')
        facility_id = request.form.get('facility_id')
        projector_id = request.form.get('projector_id')
        course_code = request.form.get('course_code')
        course_name = request.form.get('course_name')
        if course_name:
            course_name = format_course_name(course_name)
        professor = request.form.get('professor')
        
        req_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        req_start = datetime.strptime(start_str, '%H:%M').time()
        req_end = datetime.strptime(end_str, '%H:%M').time()
        
        if req_start >= req_end:
            flash("End time must be later than start time.", "danger")
            return redirect(url_for('student.reserve'))
            
        if req_start < time(7, 0) or req_end > time(21, 0):
            flash("Reservations must fall within operating hours (7:00 AM - 9:00 PM).", "danger")
            return redirect(url_for('student.reserve'))
            
        if req_date < datetime.now().date():
            flash("Cannot book reservations in the past.", "danger")
            return redirect(url_for('student.reserve'))
            
        fac_id = int(facility_id) if facility_id else None
        proj_id = int(projector_id) if projector_id else None

        if not fac_id and not proj_id:
            flash("Please select an available room or projector before submitting.", "danger")
            return redirect(url_for('student.reserve'))
        
        conn = get_db_connection()
        cursor = conn.cursor()
        conflict = has_reservation_conflict(cursor, req_date, req_start, req_end, fac_id, proj_id)
        conn.close()
                        
        if conflict:
            # Generate AI alternative suggestions
            alternatives = []
            ai_settings = load_ai_settings()
            
            # Alternative 1: Find another available room of the same type in the same slot
            if ai_settings['auto_suggest'] and fac_id:
                type_rooms = [f for f in FACILITIES_DB if f['type'] == facility_type and f['id'] != fac_id and f['status'] not in FACILITY_HARD_BLOCK_STATUSES]
                for room in type_rooms:
                    room_conflict = False
                    for res in RESERVATIONS_DB:
                        if res['status'] in ['APPROVED', 'PENDING APPROVAL', 'IN USE'] and res['facility_id'] == room['id']:
                            if res['schedule_date'] == req_date and req_start < res['end_time'] and req_end > res['start_time']:
                                room_conflict = True
                                break
                    if not room_conflict:
                        alternatives.append({
                            'type': 'alternative_facility',
                            'message': f"Room {room['code']} is available during your requested slot.",
                            'data': {
                                'facility_id': room['id'],
                                'projector_id': '',
                                'schedule_date': date_str,
                                'start_time': start_str,
                                'end_time': end_str
                            }
                        })
                        break
                        
            # Alternative 2: Offer a later/earlier slot for the same resource on the same day
            # Let's shift by 2 hours later
            later_start = (datetime.combine(req_date, req_start) + timedelta(hours=2)).time()
            later_end = (datetime.combine(req_date, req_end) + timedelta(hours=2)).time()
            
            if ai_settings['auto_suggest'] and later_end <= time(21, 0): # Operating hours check
                slot_conflict = False
                if fac_id:
                    for res in RESERVATIONS_DB:
                        if res['status'] in ['APPROVED', 'PENDING APPROVAL', 'IN USE'] and res['facility_id'] == fac_id:
                            if res['schedule_date'] == req_date and later_start < res['end_time'] and later_end > res['start_time']:
                                slot_conflict = True
                                break
                elif proj_id:
                    for res in RESERVATIONS_DB:
                        if res['status'] in ['APPROVED', 'PENDING APPROVAL'] and res['projector_id'] == proj_id:
                            if res['schedule_date'] == req_date and later_start < res['end_time'] and later_end > res['start_time']:
                                slot_conflict = True
                                break
                                
                if not slot_conflict:
                    alternatives.append({
                        'type': 'alternative_slot',
                        'message': f"Same resource is available later on the same day: {later_start.strftime('%I:%M %p')} - {later_end.strftime('%I:%M %p')}.",
                        'data': {
                            'facility_id': fac_id or '',
                            'projector_id': proj_id or '',
                            'schedule_date': date_str,
                            'start_time': later_start.strftime('%H:%M'),
                            'end_time': later_end.strftime('%H:%M')
                        }
                    })

            # Alternative 3: Offer a slot on the next day
            next_date = req_date + timedelta(days=1)
            next_conflict = False
            if fac_id:
                for res in RESERVATIONS_DB:
                    if res['status'] in ['APPROVED', 'PENDING APPROVAL', 'IN USE'] and res['facility_id'] == fac_id:
                        if res['schedule_date'] == next_date and req_start < res['end_time'] and req_end > res['start_time']:
                            next_conflict = True
                            break
            elif proj_id:
                for res in RESERVATIONS_DB:
                    if res['status'] in ['APPROVED', 'PENDING APPROVAL'] and res['projector_id'] == proj_id:
                        if res['schedule_date'] == next_date and req_start < res['end_time'] and req_end > res['start_time']:
                            next_conflict = True
                            break
                            
            if ai_settings['auto_suggest'] and not next_conflict:
                alternatives.append({
                    'type': 'alternative_slot',
                    'message': f"Same slot is fully available tomorrow: {next_date.strftime('%b %d, %Y')} at {req_start.strftime('%I:%M %p')}.",
                    'data': {
                        'facility_id': fac_id or '',
                        'projector_id': proj_id or '',
                        'schedule_date': next_date.strftime('%Y-%m-%d'),
                        'start_time': start_str,
                        'end_time': end_str
                    }
                })

            return render_template('student/reservation_conflict.html', 
                                   date=req_date.strftime('%b %d, %Y'), 
                                   start=req_start.strftime('%I:%M %p'), 
                                   end=req_end.strftime('%I:%M %p'), 
                                   facility_type=facility_type,
                                   alternatives=alternatives)
                                   
        # No conflict: Save reservation while holding a short resource lock.
        conn = get_db_connection()
        cursor = conn.cursor()
        lock_acquired = False
        try:
            lock_acquired = _acquire_reservation_lock(cursor, fac_id, proj_id, req_date.strftime('%Y-%m-%d'))
            if not lock_acquired:
                flash("That resource is busy right now. Please try submitting again.", "warning")
                return redirect(url_for('student.reserve'))

            if has_reservation_conflict(cursor, req_date, req_start, req_end, fac_id, proj_id):
                conn.rollback()
                flash("That time slot has just been taken. Please choose another schedule.", "warning")
                return redirect(url_for('student.reserve'))

            cursor.execute('''
                INSERT INTO reservation_requests (student_id, facility_type, facility_id, projector_id, schedule_date, start_time, end_time, course_code, course_name, professor, status, remarks)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 'PENDING APPROVAL', NULL)
            ''', (u.id, facility_type, fac_id or None, proj_id or None, req_date.strftime('%Y-%m-%d'), req_start.strftime('%H:%M:%S'), req_end.strftime('%H:%M:%S'), course_code, course_name, professor))
            conn.commit()
            new_id = cursor.lastrowid
        finally:
            if lock_acquired:
                _release_reservation_lock(cursor, fac_id, proj_id, req_date.strftime('%Y-%m-%d'))
            conn.close()
        
        sync_db_to_memory()
        
        log_event(u.id, "CREATE RESERVATION", f"Created pending reservation request #{new_id} for {facility_type}.")
        flash("Reservation request submitted successfully!", "success")
        return redirect(url_for('student.dashboard'))
        
    return render_template('student/reservation.html')

@student_bp.route('/student/cancel/<int:id>', methods=['POST'])
def cancel(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'student':
        return redirect(url_for('auth.login'))
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT facility_id, projector_id, status FROM reservation_requests WHERE id = ? AND student_id = ?", (id, u.id))
    row = cursor.fetchone()
    if not row:
        conn.close()
        flash("Reservation request not found.", "danger")
        return redirect(request.referrer or url_for('student.dashboard'))
    if row['status'] != 'PENDING APPROVAL':
        conn.close()
        flash("Only pending reservation requests can be cancelled from the student portal.", "warning")
        return redirect(request.referrer or url_for('student.dashboard'))

    cursor.execute("UPDATE reservation_requests SET status = 'CANCELLED' WHERE id = ? AND student_id = ?", (id, u.id))
    _release_facility_if_free(cursor, row['facility_id'], id)
    _release_projector_if_free(cursor, row['projector_id'], id)
    conn.commit()
    conn.close()
    
    sync_db_to_memory()
    
    log_event(u.id, "CANCEL RESERVATION", f"Cancelled reservation request #{id}.")
    flash("Reservation request cancelled.", "info")
    return redirect(request.referrer or url_for('student.dashboard'))

@student_bp.route('/student/profile', methods=['GET', 'POST'])
def profile():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'student':
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        contact_number = request.form.get('contact_number')
        program = request.form.get('program')
        year_section = request.form.get('year_section')
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_new_password = request.form.get('confirm_new_password', '').strip()
        
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT password_hash FROM authorized_users WHERE id = ?", (u.id,))
        row = cursor.fetchone()
        
        if new_password:
            if not row or not check_password_hash(row['password_hash'], current_password):
                conn.close()
                return render_template('student/profile.html', form=get_mock_form(), current_password_error="Current password is incorrect.")
                
            val = new_password
            has_length = len(val) >= 8
            has_upper = any(c.isupper() for c in val)
            has_lower = any(c.islower() for c in val)
            has_number = any(c.isdigit() for c in val)
            has_special = any(not c.isalnum() for c in val)
            
            if not (has_length and has_upper and has_lower and has_number and has_special):
                conn.close()
                flash("New password does not meet requirements. It must have at least 8 characters, 1 uppercase letter, 1 lowercase letter, 1 number, and 1 special character.", "danger")
                return render_template('student/profile.html', form=get_mock_form())

            if check_password_hash(row['password_hash'], new_password):
                conn.close()
                return render_template('student/profile.html', form=get_mock_form(), new_password_error="New password cannot be the same as your current password.")
                
            if new_password != confirm_new_password:
                conn.close()
                flash("Passwords should match.", "danger")
                return render_template('student/profile.html', form=get_mock_form())
                
            cursor.execute('''
                UPDATE authorized_users 
                SET password_hash = ?, contact_number = ?, program = ?, year_section = ?, password_changed = 1, email_verified = 1
                WHERE id = ?
            ''', (generate_password_hash(new_password), contact_number, program, year_section, u.id))
        else:
            cursor.execute('''
                UPDATE authorized_users 
                SET contact_number = ?, program = ?, year_section = ?
                WHERE id = ?
            ''', (contact_number, program, year_section, u.id))
            
        conn.commit()
        conn.close()
        
        sync_db_to_memory()
        
        log_event(u.id, "UPDATE PROFILE", "Student updated profile credentials.")
        flash("Profile changes saved successfully!", "success")
        return redirect(url_for('student.profile'))
            
    return render_template('student/profile.html', form=get_mock_form())

@student_bp.route('/student/status')
def status():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'student':
        return redirect(url_for('auth.login'))

    refresh_facility_usage_statuses()

    facility_type = request.args.get('facility_type', '')
    date_str = request.args.get('date', '')
    status = request.args.get('status', '')
    
    student_res = [r for r in RESERVATIONS_DB if r['user_id'] == u.id]
    
    filtered = []
    for res in student_res:
        if facility_type and res['facility_type'] != facility_type:
            continue
        if date_str and str(res['schedule_date']) != date_str:
            continue
        if status and res['status'] != status:
            continue
        filtered.append(ReservationObj(res))
        
    return render_template('student/status.html', reservations=filtered, 
                           selected_type=facility_type, 
                           selected_date=date_str, 
                           selected_status=status)

@student_bp.route('/student/history')
def history():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'student':
        return redirect(url_for('auth.login'))
        
    search = request.args.get('search', '')
    date_str = request.args.get('date', '')
    
    student_res = [r for r in RESERVATIONS_DB if r['user_id'] == u.id]
    
    filtered = []
    for res in student_res:
        if date_str and str(res['schedule_date']) != date_str:
            continue
        if search:
            term = search.lower()
            cc = (res['course_code'] or '').lower()
            cn = (res['course_name'] or '').lower()
            prof = (res['professor'] or '').lower()
            if term not in cc and term not in cn and term not in prof:
                continue
        filtered.append(ReservationObj(res))
        
    return render_template('student/history.html', reservations=filtered, search=search, selected_date=date_str)

@student_bp.route('/student/export-history')
def export_history():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'student':
        return redirect(url_for('auth.login'))

    format_type = request.args.get('format', 'csv')
    student_res = [r for r in RESERVATIONS_DB if r['user_id'] == u.id]
    filtered_res = [ReservationObj(r) for r in student_res]
    student_name_safe = (u.full_name or u.email or 'user').replace(',', '').replace(' ', '_')
    base_filename = 'PUPPQ_ReservationSystem_History'

    if format_type == 'excel':
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Reservation History'

        header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid')
        align_center = Alignment(horizontal='center', vertical='center')

        # Metadata rows
        ws.append(['PUP Parañaque Resource Reservation System'])
        ws['A1'].font = Font(name='Segoe UI', size=13, bold=True, color='800000')
        ws.append(['Reservation History'])
        ws.append([])
        ws.append(['Student:', u.full_name])
        ws.append(['Student Number:', u.student_number or 'N/A'])
        ws.append(['Program, Year & Section:', f"{u.program} {u.year_section}"])
        ws.append(['Generated On:', datetime.now().strftime('%B %d, %Y %I:%M %p')])
        ws.append([])

        headers = ['ID', 'Facility Category', 'Resource Code', 'Schedule Date', 'Start Time', 'End Time', 'Course', 'Professor', 'Status', 'Remarks']
        ws.append(headers)
        header_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        for r in filtered_res:
            code = r.facility.code if r.facility else (r.projector.code if r.projector else 'N/A')
            course = f"{r.course_code} - {r.course_name}" if r.course_code else 'N/A'
            ws.append([r.id, r.facility_type, code, r.schedule_date.strftime('%Y-%m-%d'), r.start_time.strftime('%I:%M %p'), r.end_time.strftime('%I:%M %p'), course, r.professor or 'N/A', r.status, r.remarks or ''])

        for col in ws.columns:
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        response = Response(out.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response.headers['Content-Disposition'] = f'attachment; filename={base_filename}.xlsx'
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        return response

    elif format_type == 'pdf':
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors as rl_colors
        from io import BytesIO

        out = BytesIO()
        doc = SimpleDocTemplate(out, pagesize=landscape(letter), leftMargin=36, rightMargin=36, topMargin=54, bottomMargin=54)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('RTitle', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=18, leading=22, textColor=rl_colors.HexColor('#800000'), spaceAfter=4)
        meta_style = ParagraphStyle('RMeta', parent=styles['Normal'], fontName='Helvetica', fontSize=9, leading=13, textColor=rl_colors.HexColor('#475569'))
        cell_style = ParagraphStyle('RCell', parent=styles['Normal'], fontName='Helvetica', fontSize=8, leading=10, textColor=rl_colors.HexColor('#1E293B'))
        cell_bold = ParagraphStyle('RCellBold', parent=cell_style, fontName='Helvetica-Bold')
        cell_hdr = ParagraphStyle('RHdr', parent=cell_style, fontName='Helvetica-Bold', textColor=rl_colors.white, alignment=1)

        story = []
        story.append(Paragraph('RESOURCE RESERVATION HISTORY', title_style))
        story.append(Paragraph(f'<b>Student:</b> {u.full_name} &nbsp;&nbsp; <b>Student No.:</b> {u.student_number or "N/A"} &nbsp;&nbsp; <b>Program, Year & Section:</b> {u.program} {u.year_section}<br/><b>Generated On:</b> {datetime.now().strftime("%B %d, %Y %I:%M %p")} &nbsp;&nbsp; <b>Total Records:</b> {len(filtered_res)}', meta_style))
        story.append(Spacer(1, 12))

        table_data = []
        headers = ['ID', 'Category', 'Resource', 'Date', 'Time', 'Course', 'Professor', 'Status', 'Remarks']
        table_data.append([Paragraph(h, cell_hdr) for h in headers])

        STATUS_PDF_COLORS = {'APPROVED': '#059669', 'PENDING APPROVAL': '#D97706', 'REJECTED': '#DC2626', 'CANCELLED': '#6B7280', 'COMPLETED': '#0284C7'}
        for r in filtered_res:
            code = r.facility.code if r.facility else (r.projector.code if r.projector else 'N/A')
            course = f"{r.course_code}<br/><font size='7' color='#64748B'>{r.course_name}</font>" if r.course_code else 'N/A'
            times_str = f"{r.start_time.strftime('%I:%M %p')} - {r.end_time.strftime('%I:%M %p')}"
            sc = STATUS_PDF_COLORS.get(r.status, '#475569')
            status_html = f"<b><font color='{sc}'>{r.status}</font></b>"
            table_data.append([
                Paragraph(f"#{r.id}", cell_bold),
                Paragraph(r.facility_type, cell_style),
                Paragraph(code, cell_bold),
                Paragraph(r.schedule_date.strftime('%Y-%m-%d'), cell_style),
                Paragraph(times_str, cell_style),
                Paragraph(course, cell_style),
                Paragraph(r.professor or 'N/A', cell_style),
                Paragraph(status_html, cell_style),
                Paragraph(r.remarks or '', cell_style)
            ])

        col_widths = [20, 75, 55, 65, 90, 110, 100, 85, 80]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), rl_colors.HexColor('#800000')),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('GRID', (0, 0), (-1, -1), 0.5, rl_colors.HexColor('#CBD5E1')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ])
        for i in range(1, len(table_data)):
            bg = rl_colors.HexColor('#F8FAFC') if i % 2 == 1 else rl_colors.white
            t_style.add('BACKGROUND', (0, i), (-1, i), bg)
        t.setStyle(t_style)
        story.append(t)

        doc.build(story)
        out.seek(0)
        response = Response(out.getvalue(), mimetype='application/pdf')
        response.headers['Content-Disposition'] = f'attachment; filename={base_filename}.pdf'
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        return response

    else:
        # Default: CSV with metadata header
        si = StringIO()
        cw = csv.writer(si)
        cw.writerow(['PUP Parañaque Resource Reservation System'])
        cw.writerow(['Reservation History'])
        cw.writerow([])
        cw.writerow(['Student:', u.full_name])
        cw.writerow(['Student Number:', u.student_number or 'N/A'])
        cw.writerow(['Program, Year & Section:', f"{u.program} {u.year_section}"])
        cw.writerow(['Generated On:', datetime.now().strftime('%B %d, %Y %I:%M %p')])
        cw.writerow(['Total Records:', len(filtered_res)])
        cw.writerow([])
        cw.writerow(['ID', 'Facility Category', 'Resource Code', 'Schedule Date', 'Start Time', 'End Time', 'Course', 'Professor', 'Status', 'Remarks'])
        for r in filtered_res:
            code = r.facility.code if r.facility else (r.projector.code if r.projector else 'N/A')
            course = f"{r.course_code} - {r.course_name}" if r.course_code else 'N/A'
            cw.writerow([r.id, r.facility_type, code, r.schedule_date.strftime('%B %d, %Y'), r.start_time.strftime('%I:%M %p'), r.end_time.strftime('%I:%M %p'), course, r.professor or 'N/A', r.status, r.remarks or ''])
        response = Response(si.getvalue(), mimetype='text/csv')
        response.headers['Content-Disposition'] = f'attachment; filename={base_filename}.csv'
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        return response


@student_bp.route('/student/export-pdf')
def export_pdf():
    """Legacy redirect to the new multi-format export route."""
    return redirect(url_for('student.export_history', format='pdf'))


# ==========================================
# Admin Routes
# ==========================================
@admin_bp.route('/admin/dashboard')
def dashboard():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))

    refresh_facility_usage_statuses()

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT COUNT(*) FROM reservation_requests")
    total = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM reservation_requests WHERE status = 'PENDING APPROVAL'")
    pending = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM reservation_requests WHERE status = 'APPROVED'")
    approved = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM authorized_users WHERE account_status = 'ACTIVE' AND role = 'STUDENT'")
    active_users = cursor.fetchone()[0]

    cursor.execute('''
        SELECT
            r.id, r.facility_type, r.schedule_date, r.start_time, r.end_time, r.status,
            au.first_name, au.middle_name, au.last_name, au.student_number, au.program,
            f.code AS facility_code, p.code AS projector_code
        FROM reservation_requests r
        JOIN authorized_users au ON r.student_id = au.id
        LEFT JOIN facilities f ON r.facility_id = f.id
        LEFT JOIN projectors p ON r.projector_id = p.id
        WHERE r.status = 'PENDING APPROVAL'
        ORDER BY r.created_at DESC
        LIMIT 5
    ''')
    recent = []
    for row in cursor.fetchall():
        middle_name = row['middle_name'] or ''
        student_name = f"{row['last_name']}, {row['first_name']}"
        if middle_name.strip():
            student_name = f"{student_name} {middle_name.strip()[0].upper()}."

        schedule_date = row['schedule_date']
        if isinstance(schedule_date, str):
            schedule_date = datetime.strptime(schedule_date, '%Y-%m-%d').date()
        start_time = row['start_time']
        if isinstance(start_time, str):
            start_time = datetime.strptime(start_time, '%H:%M:%S' if start_time.count(':') == 2 else '%H:%M').time()
        end_time = row['end_time']
        if isinstance(end_time, str):
            end_time = datetime.strptime(end_time, '%H:%M:%S' if end_time.count(':') == 2 else '%H:%M').time()

        recent.append({
            'id': row['id'],
            'student_name': student_name,
            'student_number': row['student_number'],
            'program': row['program'] or 'N/A',
            'resource_code': row['facility_code'] or row['projector_code'] or row['facility_type'],
            'schedule_date': schedule_date,
            'start_time': start_time,
            'end_time': end_time,
            'status': row['status']
        })

    facility_counts = {
        'Audio-Visual Room (AVR)': 0,
        'Computer Laboratory': 0,
        'Hospitality Management Laboratory': 0,
        'Projector': 0
    }
    cursor.execute('''
        SELECT facility_type, COUNT(*) AS total
        FROM reservation_requests
        GROUP BY facility_type
    ''')
    for row in cursor.fetchall():
        rtype = row['facility_type']
        if rtype == 'Projector (Equipment Only)':
            rtype = 'Projector'
        if rtype in facility_counts:
            facility_counts[rtype] += row['total']
        elif rtype:
            facility_counts[rtype] = row['total']
        
    hours_occupancy = {str(h): 0 for h in range(7, 22)}
    dashboard_day_counts = Counter()
    active_booking_count = 0
    cursor.execute('''
        SELECT schedule_date, start_time, end_time
        FROM reservation_requests
        WHERE status IN ('APPROVED', 'PENDING APPROVAL', 'IN USE')
    ''')
    for row in cursor.fetchall():
        try:
            active_booking_count += 1
            sd = row['schedule_date']
            if not hasattr(sd, 'weekday'):
                sd = datetime.strptime(str(sd), '%Y-%m-%d').date()
            dashboard_day_counts[sd.weekday()] += 1
            start_hour = row['start_time'].hour if hasattr(row['start_time'], 'hour') else int(str(row['start_time']).split(':')[0])
            end_hour = row['end_time'].hour if hasattr(row['end_time'], 'hour') else int(str(row['end_time']).split(':')[0])
            for hr in range(start_hour, end_hour):
                if str(hr) in hours_occupancy:
                    hours_occupancy[str(hr)] += 1
        except Exception:
            pass
                
    weekday_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
    peak_day_idx = dashboard_day_counts.most_common(1)[0][0] if dashboard_day_counts else datetime.now().weekday()
    peak_day_val = weekday_names[peak_day_idx]

    peak_hour_val = "No active peak yet"
    max_count = 0
    for h, cnt in hours_occupancy.items():
        if cnt > max_count:
            max_count = cnt
            hr_int = int(h)
            peak_hour_val = f"{hr_int % 12 or 12}:00 {'AM' if hr_int < 12 else 'PM'} - {(hr_int + 1) % 12 or 12}:00 {'AM' if (hr_int + 1) < 12 else 'PM'}"

    if active_booking_count == 0:
        utilization_summary = "No active reservation pattern is available yet. The model will update as bookings are approved or submitted."
    elif max_count >= 5 or pending >= 5:
        utilization_summary = f"{peak_day_val} currently carries the highest demand, with {max_count} active booking(s) in the busiest hour. Review pending requests before adding more peak-hour bookings."
    else:
        utilization_summary = f"{peak_day_val} is currently the busiest day, but active booking pressure remains manageable."

    # Compute real monthly counts for the current year
    monthly_counts = [0] * 12
    current_year = datetime.now().year
    cursor.execute('''
        SELECT schedule_date, COUNT(*) AS total
        FROM reservation_requests
        WHERE schedule_date >= ? AND schedule_date < ?
        GROUP BY schedule_date
    ''', (f'{current_year}-01-01', f'{current_year + 1}-01-01'))
    for row in cursor.fetchall():
        try:
            sd = row['schedule_date']
            if hasattr(sd, 'month'):
                month_idx = sd.month - 1
            else:
                month_idx = datetime.strptime(str(sd), '%Y-%m-%d').month - 1
            if 0 <= month_idx <= 11:
                monthly_counts[month_idx] += row['total']
        except Exception:
            pass

    conn.close()

    insights = {
        'peak_hour': peak_hour_val,
        'forecast_alert': utilization_summary,
        'peak_day': peak_day_val,
        'facility_utilization': facility_counts,
        'hours_occupancy': hours_occupancy,
        'monthly_counts': monthly_counts
    }
    
    return render_template('admin/dashboard.html', total=total, pending=pending, approved=approved, active_users=active_users, recent=recent, insights=insights)

@admin_bp.route('/admin/requests')
def requests():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))

    refresh_facility_usage_statuses()
    refresh_projector_custody_statuses()
        
    active_statuses = {'APPROVED', 'IN USE', 'AWAITING RETURN', 'OVERDUE'}
    all_requests = [ReservationObj(r) for r in sorted(RESERVATIONS_DB, key=lambda item: item['created_at'], reverse=True)]
    pending = [ReservationObj(r) for r in RESERVATIONS_DB if r['status'] == 'PENDING APPROVAL']
    active = [ReservationObj(r) for r in RESERVATIONS_DB if r['status'] in active_statuses]
    processed = [ReservationObj(r) for r in RESERVATIONS_DB if r['status'] not in active_statuses and r['status'] != 'PENDING APPROVAL']
    
    return render_template('admin/requests.html', all_requests=all_requests, pending=pending, active=active, processed=processed)

@admin_bp.route('/admin/requests/approve/<int:id>', methods=['POST'])
def requests_approve(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    remarks = request.form.get('remarks', '')
    
    # Get reservation and student details for email
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT r.facility_type, r.facility_id, r.projector_id, r.schedule_date, r.start_time, r.end_time, r.status, u.pup_email, u.first_name, u.last_name
        FROM reservation_requests r
        JOIN authorized_users u ON r.student_id = u.id
        WHERE r.id = ?
    ''', (id,))
    res_row = cursor.fetchone()

    if not res_row:
        conn.close()
        flash("Reservation request not found.", "danger")
        return redirect(url_for('admin.requests'))

    if res_row['status'] != 'PENDING APPROVAL':
        conn.close()
        flash("Only pending reservation requests can be approved.", "warning")
        return redirect(url_for('admin.requests'))

    lock_acquired = False
    try:
        lock_acquired = _acquire_reservation_lock(cursor, res_row['facility_id'], res_row['projector_id'], res_row['schedule_date'])
        if not lock_acquired:
            flash("That resource is busy right now. Please try approving again.", "warning")
            return redirect(url_for('admin.requests'))

        if has_reservation_conflict(
            cursor,
            res_row['schedule_date'],
            res_row['start_time'],
            res_row['end_time'],
            res_row['facility_id'],
            res_row['projector_id'],
            statuses=('APPROVED', 'IN USE'),
            exclude_id=id
        ):
            conn.rollback()
            flash("This request now conflicts with an approved reservation and cannot be approved.", "danger")
            return redirect(url_for('admin.requests'))

        if res_row['projector_id']:
            cursor.execute("SELECT status FROM projectors WHERE id = ?", (res_row['projector_id'],))
            projector_row = cursor.fetchone()
            if projector_row and projector_row['status'] in PROJECTOR_HARD_BLOCK_STATUSES:
                conn.rollback()
                flash("This projector cannot be approved because it is checked out, unavailable, or under maintenance.", "danger")
                return redirect(url_for('admin.requests'))
        elif res_row['facility_id']:
            cursor.execute("SELECT status FROM facilities WHERE id = ?", (res_row['facility_id'],))
            facility_row = cursor.fetchone()
            if facility_row and facility_row['status'] in FACILITY_HARD_BLOCK_STATUSES:
                conn.rollback()
                flash("This facility cannot be approved because it is unavailable or under maintenance.", "danger")
                return redirect(url_for('admin.requests'))

        cursor.execute("UPDATE reservation_requests SET status = 'APPROVED', remarks = ? WHERE id = ?", (remarks, id))
        if res_row and res_row['projector_id']:
            cursor.execute("UPDATE projectors SET status = 'Reserved' WHERE id = ?", (res_row['projector_id'],))
        elif res_row and res_row['facility_id']:
            cursor.execute("UPDATE facilities SET status = 'Reserved' WHERE id = ?", (res_row['facility_id'],))
        conn.commit()
    finally:
        if lock_acquired:
            _release_reservation_lock(cursor, res_row['facility_id'], res_row['projector_id'], res_row['schedule_date'])
        conn.close()

    sync_db_to_memory()
    log_event(u.id, "APPROVE RESERVATION", f"Approved reservation request #{id}.")

    student_email = res_row['pup_email']
    student_name = f"{res_row['first_name']} {res_row['last_name']}"
    resource_type = res_row['facility_type']
    code = _get_reservation_resource_code(res_row)
    s_date_str, s_time_str, e_time_str = _format_reservation_schedule_for_email(res_row)

    subject = f"PUP Reservation System: Reservation Request #{id} APPROVED"
    body = f"Dear {student_name},\n\nYour reservation request #{id} has been APPROVED.\n\nDetails:\n- Resource: {resource_type} ({code})\n- Date: {s_date_str}\n- Time: {s_time_str} - {e_time_str}\n- Administrative Remarks: {remarks or 'N/A'}\n\nThank you,\nPUP Paranaque Administration"
    send_notification_email_async(student_email, subject, body)

    flash(f"Reservation #{id} has been approved.", "success")
    return redirect(url_for('admin.requests'))

@admin_bp.route('/admin/requests/reject/<int:id>', methods=['POST'])
def requests_reject(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    remarks = request.form.get('remarks', '')
    if not remarks:
        flash("Rejection reason is required.", "danger")
        return redirect(url_for('admin.requests'))
        
    # Get reservation and student details for email
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT r.facility_type, r.facility_id, r.projector_id, r.schedule_date, r.start_time, r.end_time, r.status, u.pup_email, u.first_name, u.last_name
        FROM reservation_requests r
        JOIN authorized_users u ON r.student_id = u.id
        WHERE r.id = ?
    ''', (id,))
    res_row = cursor.fetchone()

    if not res_row:
        conn.close()
        flash("Reservation request not found.", "danger")
        return redirect(url_for('admin.requests'))

    if res_row['status'] != 'PENDING APPROVAL':
        conn.close()
        flash("Only pending reservation requests can be rejected.", "warning")
        return redirect(url_for('admin.requests'))
    
    cursor.execute("UPDATE reservation_requests SET status = 'REJECTED', remarks = ? WHERE id = ?", (remarks, id))
    conn.commit()
    conn.close()
    
    sync_db_to_memory()
    
    log_event(u.id, "REJECT RESERVATION", f"Rejected reservation request #{id} with reason: '{remarks}'.")
    
    # Send email notification asynchronously
    if res_row:
        student_email = res_row['pup_email']
        student_name = f"{res_row['first_name']} {res_row['last_name']}"
        resource_type = res_row['facility_type']
        
        # Get code
        code = 'N/A'
        conn = get_db_connection()
        cursor = conn.cursor()
        if res_row['facility_id']:
            cursor.execute("SELECT code FROM facilities WHERE id = ?", (res_row['facility_id'],))
            row = cursor.fetchone()
            if row: code = row['code']
        elif res_row['projector_id']:
            cursor.execute("SELECT code FROM projectors WHERE id = ?", (res_row['projector_id'],))
            row = cursor.fetchone()
            if row: code = row['code']
        conn.close()
        
        # Format times
        from datetime import datetime
        try:
            s_date = res_row['schedule_date']
            if isinstance(s_date, str):
                s_date_obj = datetime.strptime(s_date, '%Y-%m-%d')
                s_date_str = s_date_obj.strftime('%B %d, %Y')
            else:
                s_date_str = s_date.strftime('%B %d, %Y')
                
            s_time = res_row['start_time']
            if isinstance(s_time, str):
                s_time_obj = datetime.strptime(s_time, '%H:%M:%S' if s_time.count(':')==2 else '%H:%M')
                s_time_str = s_time_obj.strftime('%I:%M %p')
            else:
                s_time_str = s_time.strftime('%I:%M %p')
                
            e_time = res_row['end_time']
            if isinstance(e_time, str):
                e_time_obj = datetime.strptime(e_time, '%H:%M:%S' if e_time.count(':')==2 else '%H:%M')
                e_time_str = e_time_obj.strftime('%I:%M %p')
            else:
                e_time_str = e_time.strftime('%I:%M %p')
        except Exception as e:
            print("Error parsing datetime for email:", e)
            s_date_str = str(res_row['schedule_date'])
            s_time_str = str(res_row['start_time'])
            e_time_str = str(res_row['end_time'])
            
        subject = f"PUP Reservation System: Reservation Request #{id} REJECTED"
        body = f"Dear {student_name},\n\nWe regret to inform you that your reservation request #{id} has been REJECTED.\n\nDetails:\n- Resource: {resource_type} ({code})\n- Date: {s_date_str}\n- Time: {s_time_str} - {e_time_str}\n- Reason for Rejection: {remarks}\n\nYou can submit a new reservation request with updated parameters on the student portal.\n\nThank you,\nPUP Parañaque Administration"
        
        send_notification_email_async(student_email, subject, body)
        
    flash(f"Reservation #{id} has been rejected.", "danger")
    return redirect(url_for('admin.requests'))

@admin_bp.route('/admin/requests/checkout/<int:id>', methods=['POST'])
def requests_checkout(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))

    received_by = request.form.get('received_by', '').strip()
    if not received_by:
        flash("Received by is required before releasing a projector.", "danger")
        return redirect(url_for('admin.requests'))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, projector_id, status, checkout_time, return_time
        FROM reservation_requests
        WHERE id = ?
    ''', (id,))
    res_row = cursor.fetchone()

    if not res_row or not res_row['projector_id']:
        conn.close()
        flash("Only projector reservations can be checked out.", "danger")
        return redirect(url_for('admin.requests'))
    if res_row['status'] != 'APPROVED' or res_row['checkout_time'] or res_row['return_time']:
        conn.close()
        flash("Only approved, unclaimed projector reservations can be checked out.", "warning")
        return redirect(url_for('admin.requests'))

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute('''
        UPDATE reservation_requests
        SET checkout_time = ?, released_by = ?, received_by = ?
        WHERE id = ?
    ''', (now_str, u.id, received_by, id))
    cursor.execute("UPDATE projectors SET status = 'Checked Out' WHERE id = ?", (res_row['projector_id'],))
    conn.commit()
    conn.close()

    sync_db_to_memory()
    log_event(u.id, "PROJECTOR CHECKOUT", f"Released projector for reservation #{id} to {received_by}.")
    flash(f"Projector for reservation #{id} marked as checked out.", "success")
    return redirect(url_for('admin.requests'))

@admin_bp.route('/admin/requests/return/<int:id>', methods=['POST'])
def requests_return(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))

    condition_notes = request.form.get('equipment_condition', '').strip()

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, projector_id, status, checkout_time, return_time
        FROM reservation_requests
        WHERE id = ?
    ''', (id,))
    res_row = cursor.fetchone()

    if not res_row or not res_row['projector_id']:
        conn.close()
        flash("Only projector reservations can be returned.", "danger")
        return redirect(url_for('admin.requests'))
    if res_row['return_time']:
        conn.close()
        flash("This projector reservation has already been returned.", "warning")
        return redirect(url_for('admin.requests'))
    if res_row['status'] not in ('APPROVED', 'AWAITING RETURN', 'OVERDUE'):
        conn.close()
        flash("Only active projector reservations can be marked returned.", "warning")
        return redirect(url_for('admin.requests'))

    now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    cursor.execute('''
        UPDATE reservation_requests
        SET status = 'COMPLETED',
            return_time = ?,
            returned_to = ?,
            equipment_condition = ?
        WHERE id = ?
    ''', (now_str, u.id, condition_notes, id))
    _release_projector_if_free(cursor, res_row['projector_id'], id)
    conn.commit()
    conn.close()

    sync_db_to_memory()
    log_event(u.id, "PROJECTOR RETURN", f"Marked projector reservation #{id} as returned. Condition: {condition_notes or 'No notes'}.")
    flash(f"Projector for reservation #{id} returned and reservation completed.", "success")
    return redirect(url_for('admin.requests'))

@admin_bp.route('/admin/requests/complete-facility/<int:id>', methods=['POST'])
def requests_complete_facility(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, facility_id, projector_id, status
        FROM reservation_requests
        WHERE id = ?
    ''', (id,))
    res_row = cursor.fetchone()

    if not res_row or not res_row['facility_id'] or res_row['projector_id']:
        conn.close()
        flash("Only facility reservations can be marked completed here.", "danger")
        return redirect(url_for('admin.requests'))
    if res_row['status'] not in ('APPROVED', 'IN USE'):
        conn.close()
        flash("Only active facility reservations can be marked completed.", "warning")
        return redirect(url_for('admin.requests'))

    cursor.execute("UPDATE reservation_requests SET status = 'COMPLETED' WHERE id = ?", (id,))
    _sync_facility_status_from_reservations(cursor, res_row['facility_id'])
    conn.commit()
    conn.close()

    sync_db_to_memory()
    log_event(u.id, "COMPLETE FACILITY RESERVATION", f"Marked facility reservation #{id} as completed.")
    flash(f"Facility reservation #{id} marked as completed.", "success")
    return redirect(url_for('admin.requests'))

@admin_bp.route('/admin/facilities', methods=['GET', 'POST'])
def facilities():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    refresh_facility_usage_statuses()

    if request.method == 'POST':
        code = request.form.get('code')
        facility_type = request.form.get('type')
        status = request.form.get('status')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO facilities (code, type, status) VALUES (?, ?, ?)", (code, facility_type, status))
            conn.commit()
            log_event(u.id, "ADD FACILITY", f"Added facility room {code} ({facility_type}) to registry.")
            flash(f"Room {code} registered successfully.", "success")
        except Exception as e:
            import pymysql
            if isinstance(e, pymysql.err.IntegrityError):
                flash(f"Error: Room code '{code}' is already registered.", "danger")
            else:
                flash(f"Error registering room: {str(e)}", "danger")
        finally:
            conn.close()
            
        sync_db_to_memory()
        return redirect(url_for('admin.facilities'))
        
    rooms_list = [FacilityObj(f) for f in FACILITIES_DB]
    return render_template('admin/facilities.html', rooms=rooms_list, form=MockForm())

@admin_bp.route('/admin/facilities/archive/<int:id>', methods=['POST'])
def archive_facility(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT code, status FROM facilities WHERE id = ?", (id,))
    row = cursor.fetchone()
    if row:
        code = row['code']
        if row['status'] == 'Archived':
            cursor.execute("UPDATE facilities SET status = 'Available' WHERE id = ?", (id,))
            log_event(u.id, "UNARCHIVE FACILITY", f"Unarchived facility room {code}.")
            flash(f"Room {code} unarchived.", "success")
        else:
            cursor.execute("UPDATE facilities SET status = 'Archived' WHERE id = ?", (id,))
            log_event(u.id, "ARCHIVE FACILITY", f"Archived facility room {code}.")
            flash(f"Room {code} archived.", "info")
        conn.commit()
    conn.close()
    
    sync_db_to_memory()
    return redirect(url_for('admin.facilities'))

@admin_bp.route('/admin/facilities/update/<int:id>', methods=['POST'])
def update_facility_status(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))

    new_status = request.form.get('status')
    allowed_statuses = {'Available', 'Maintenance', 'Unavailable'}
    if new_status not in allowed_statuses:
        flash("Invalid facility status selected.", "danger")
        return redirect(url_for('admin.facilities'))

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT code, status FROM facilities WHERE id = ?", (id,))
    row = cursor.fetchone()
    if row:
        code = row['code']
        old_status = row['status']
        if _active_facility_reservation_exists(cursor, id):
            conn.close()
            flash(f"Facility {code} is tied to an active reservation and cannot be updated manually until it is completed.", "warning")
            return redirect(url_for('admin.facilities'))
        cursor.execute("UPDATE facilities SET status = ? WHERE id = ?", (new_status, id))
        conn.commit()
        log_event(u.id, "UPDATE FACILITY STATUS", f"Updated status of room {code} from {old_status} to {new_status}.")
        flash(f"Facility {code} status updated to {new_status}.", "success")
    conn.close()

    sync_db_to_memory()
    return redirect(url_for('admin.facilities'))

@admin_bp.route('/admin/projectors', methods=['GET', 'POST'])
def projectors():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))

    refresh_projector_custody_statuses()
        
    if request.method == 'POST':
        code = request.form.get('code')
        model = request.form.get('model')
        status = request.form.get('status')
        
        conn = get_db_connection()
        cursor = conn.cursor()
        try:
            cursor.execute("INSERT INTO projectors (code, model, status) VALUES (?, ?, ?)", (code, model, status))
            conn.commit()
            log_event(u.id, "ADD PROJECTOR", f"Added equipment unit {code} ({model}) to inventory.")
            flash(f"Projector {code} registered successfully.", "success")
        except Exception as e:
            import pymysql
            if isinstance(e, pymysql.err.IntegrityError):
                flash(f"Error: Projector code '{code}' is already registered.", "danger")
            else:
                flash(f"Error registering projector: {str(e)}", "danger")
        finally:
            conn.close()
            
        sync_db_to_memory()
        return redirect(url_for('admin.projectors'))
        
    reserved_projector_ids = _get_reserved_projector_ids()
    proj_list = []
    for p in PROJECTORS_DB:
        projector = ProjectorObj(p)
        if projector.id in reserved_projector_ids and projector.status not in ('Checked Out', 'Overdue'):
            projector.status = 'Reserved'
        proj_list.append(projector)
    return render_template('admin/projectors.html', projectors=proj_list, form=MockForm())

@admin_bp.route('/admin/projectors/update/<int:id>', methods=['POST'])
def update_projector_status(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    new_status = request.form.get('status')
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT code, status FROM projectors WHERE id = ?", (id,))
    row = cursor.fetchone()
    if row:
        code = row['code']
        old_status = row['status']
        if _active_projector_reservation_exists(cursor, id):
            conn.close()
            flash(f"Projector {code} is tied to an active reservation and cannot be updated manually until it is returned or completed.", "warning")
            return redirect(url_for('admin.projectors'))
        cursor.execute("UPDATE projectors SET status = ? WHERE id = ?", (new_status, id))
        conn.commit()
        log_event(u.id, "UPDATE PROJECTOR STATUS", f"Updated status of unit {code} from {old_status} to {new_status}.")
        flash(f"Projector {code} status updated to {new_status}.", "success")
    conn.close()
    
    sync_db_to_memory()
    return redirect(url_for('admin.projectors'))

@admin_bp.route('/admin/users')
def users():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
    return redirect(url_for('admin.students'))

@admin_bp.route('/admin/users/toggle/<int:id>', methods=['POST'])
def toggle_user_active(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT pup_email, account_status FROM authorized_users WHERE id = ?", (id,))
    row = cursor.fetchone()
    if row:
        email = row['pup_email']
        new_status = 'DISABLED' if row['account_status'] == 'ACTIVE' else 'ACTIVE'
        cursor.execute("UPDATE authorized_users SET account_status = ? WHERE id = ?", (new_status, id))
        conn.commit()
        state = "deactivated" if new_status == 'DISABLED' else "activated"
        log_event(u.id, "TOGGLE USER STATE", f"Set state of student {email} to {state}.")
        flash(f"User {email} has been {state}.", "success")
    conn.close()
    
    sync_db_to_memory()
    return redirect(request.referrer or url_for('admin.students'))

@admin_bp.route('/admin/users/reset/<int:id>', methods=['POST'])
def reset_user_password(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT pup_email FROM authorized_users WHERE id = ?", (id,))
    row = cursor.fetchone()
    if row:
        email = row['pup_email']
        cursor.execute("UPDATE authorized_users SET password_hash = ?, password_changed = 0, email_verified = 0 WHERE id = ?",
                       (generate_password_hash("PUPrs@1904"), id))
        conn.commit()
        log_event(u.id, "RESET USER PASSWORD", f"Reset password of user {email} to default.")
        flash(f"Password for {email} reset to default.", "success")
    conn.close()
    
    sync_db_to_memory()
    return redirect(request.referrer or url_for('admin.students'))

@admin_bp.route('/admin/students', methods=['GET', 'POST'])
def students():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        # CSV / Excel Masterlist Upload
        file = request.files.get('file')
        if not file or file.filename == '':
            flash("Please choose a CSV or Excel file to upload.", "warning")
            return redirect(url_for('admin.students'))

        filename = file.filename.lower()
        imported_count = 0
        skipped_count = 0

        try:
            rows = []
            if filename.endswith('.csv'):
                raw = file.stream.read()
                try:
                    text = raw.decode('utf-8-sig')
                except UnicodeDecodeError:
                    text = raw.decode('latin-1')
                stream = StringIO(text, newline=None)
                reader = csv.reader(stream)
                next(reader, None)  # skip header
                for row in reader:
                    if any(c.strip() for c in row):
                        rows.append([c.strip() for c in row])
            elif filename.endswith('.xlsx') or filename.endswith('.xls'):
                import openpyxl
                from io import BytesIO
                wb = openpyxl.load_workbook(BytesIO(file.stream.read()), data_only=True)
                ws = wb.active
                first_row = True
                for row in ws.iter_rows(values_only=True):
                    if first_row:
                        first_row = False
                        continue  # skip header row
                    row = [str(c).strip() if c is not None else '' for c in row]
                    if any(c for c in row):
                        rows.append(row)
            else:
                flash("Unsupported file type. Please upload a .csv or .xlsx file.", "danger")
                return redirect(url_for('admin.students'))

            for row in rows:
                if len(row) >= 7:
                    stud_num = row[0]
                    email = row[1]
                    last_name = row[2]
                    first_name = row[3]
                    middle_name = row[4]
                    program = row[5] or 'BSIT'
                    year_section = row[6] or '1-1'
                elif len(row) >= 4:
                    stud_num = row[0]
                    first_name = row[1]
                    last_name = row[2]
                    email = row[3]
                    middle_name = ''
                    program = 'BSIT'
                    year_section = '1-1'
                else:
                    skipped_count += 1
                    continue

                if not stud_num or not email:
                    skipped_count += 1
                    continue

                success, msg = import_student_to_masterlist(
                    stud_num, first_name, last_name, email,
                    middle_name=middle_name, program=program, year_section=year_section
                )
                if success:
                    imported_count += 1
                else:
                    skipped_count += 1

            sync_db_to_memory()
            log_event(u.id, "IMPORT MASTERLIST", f"Imported {imported_count} user accounts from masterlist file ({skipped_count} skipped/duplicate).")
            if skipped_count > 0:
                flash(f"Imported {imported_count} accounts successfully. {skipped_count} rows were skipped (duplicate or incomplete).", "warning")
            else:
                flash(f"Successfully imported {imported_count} user accounts from masterlist.", "success")

        except Exception as e:
            flash(f"Error reading file: {str(e)}", "danger")

        return redirect(url_for('admin.students'))

    stud_list = [MockUser(usr) for usr in USERS_DB if usr.get('role') != 'admin']
    return render_template('admin/students.html', students=stud_list, form=MockForm())

@admin_bp.route('/admin/students/reset/<int:id>', methods=['POST'])
def reset_student(id):
    return reset_user_password(id)

@admin_bp.route('/admin/students/reverify/<int:id>', methods=['POST'])
def force_reverify(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT pup_email FROM authorized_users WHERE id = ?", (id,))
    row = cursor.fetchone()
    if row:
        email = row['pup_email']
        cursor.execute("UPDATE authorized_users SET password_changed = 0, email_verified = 0 WHERE id = ?", (id,))
        conn.commit()
        log_event(u.id, "FORCE REVERIFY", f"Forced user {email} to re-verify credentials.")
        flash(f"Forced re-verification for {email}.", "info")
    conn.close()
    
    sync_db_to_memory()
    return redirect(url_for('admin.students'))

@admin_bp.route('/admin/students/toggle/<int:id>', methods=['POST'])
def toggle_student(id):
    return toggle_user_active(id)

@admin_bp.route('/admin/students/edit/<int:id>', methods=['POST'])
def edit_student(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    student_number = request.form.get('student_number', '').strip()
    first_name = request.form.get('first_name', '').strip()
    middle_name = request.form.get('middle_name', '').strip()
    last_name = request.form.get('last_name', '').strip()
    program = request.form.get('program', '').strip()
    year_section = request.form.get('year_section', '').strip()
    pup_email = request.form.get('pup_email', '').strip()
    
    if not all([student_number, first_name, last_name, program, year_section, pup_email]):
        flash("All fields except Middle Name are required.", "danger")
        return redirect(url_for('admin.students'))
        
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Check duplicate email or student number in authorized_users for OTHER students
        cursor.execute('''
            SELECT student_number FROM authorized_users WHERE id = ?
        ''', (id,))
        current_row = cursor.fetchone()
        if not current_row:
            flash("Student not found.", "danger")
            return redirect(url_for('admin.students'))
            
        current_stud_num = current_row['student_number']
        
        cursor.execute('''
            SELECT COUNT(*) FROM authorized_users 
            WHERE (student_number = ? AND student_number != ?)
               OR (pup_email = ? AND student_number != ?)
        ''', (student_number, current_stud_num, pup_email, current_stud_num))
        
        if cursor.fetchone()[0] > 0:
            flash("Student Number or PUP Webmail is already registered to another student.", "danger")
            return redirect(url_for('admin.students'))
            
        cursor.execute("SET FOREIGN_KEY_CHECKS = 0")
        
        # Update authorized_users
        cursor.execute('''
            UPDATE authorized_users 
            SET student_number = ?, first_name = ?, middle_name = ?, last_name = ?, program = ?, year_section = ?, pup_email = ?
            WHERE id = ?
        ''', (student_number, first_name, middle_name, last_name, program, year_section, pup_email, id))
        
        conn.commit()
        log_event(u.id, "EDIT STUDENT", f"Updated student profile for {pup_email} (ID: {id}).")
        flash("Student information updated successfully.", "success")
    except Exception as e:
        conn.rollback()
        flash(f"Error updating student: {str(e)}", "danger")
    finally:
        cursor.execute("SET FOREIGN_KEY_CHECKS = 1")
        conn.close()
        
    sync_db_to_memory()
    return redirect(url_for('admin.students'))
@admin_bp.route('/admin/students/otp-history/<int:id>')
def student_otp_history(id):
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401
        
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT expires_at, is_used, created_at 
        FROM otp_verifications 
        WHERE user_id = ?
        ORDER BY created_at DESC
    ''', (id,))
    rows = cursor.fetchall()
    conn.close()
    
    history_list = []
    for row in rows:
        expires = datetime.strptime(row['expires_at'], '%Y-%m-%d %H:%M:%S')
        is_expired = datetime.now() > expires
        
        status = "Used" if row['is_used'] else ("Expired" if is_expired else "Active")
        history_list.append({
            'otp_code': '******',
            'expires_at': row['expires_at'],
            'is_used': bool(row['is_used']),
            'created_at': row['created_at'],
            'status': status
        })
        
    return jsonify({'otp_history': history_list})

@admin_bp.route('/admin/calendar')
def calendar():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
    return render_template('admin/calendar.html')

@admin_bp.route('/admin/calendar/events')
def calendar_events():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401
        
    refresh_facility_usage_statuses()

    # Status color mapping for FullCalendar event styling
    STATUS_COLORS = {
        'APPROVED': {'backgroundColor': '#059669', 'borderColor': '#047857', 'textColor': '#ffffff'},
        'IN USE': {'backgroundColor': '#0F766E', 'borderColor': '#115E59', 'textColor': '#ffffff'},
        'PENDING APPROVAL': {'backgroundColor': '#D97706', 'borderColor': '#B45309', 'textColor': '#ffffff'},
        'REJECTED': {'backgroundColor': '#DC2626', 'borderColor': '#B91C1C', 'textColor': '#ffffff'},
        'AWAITING RETURN': {'backgroundColor': '#7C3AED', 'borderColor': '#6D28D9', 'textColor': '#ffffff'},
        'OVERDUE': {'backgroundColor': '#B91C1C', 'borderColor': '#991B1B', 'textColor': '#ffffff'},
        'CANCELLED': {'backgroundColor': '#6B7280', 'borderColor': '#4B5563', 'textColor': '#ffffff'},
        'COMPLETED': {'backgroundColor': '#0284C7', 'borderColor': '#0369A1', 'textColor': '#ffffff'},
    }

    start_arg = request.args.get('start', '')
    end_arg = request.args.get('end', '')
    start_date = start_arg[:10] if len(start_arg) >= 10 else None
    end_date = end_arg[:10] if len(end_arg) >= 10 else None

    conn = get_db_connection()
    cursor = conn.cursor()
    params = []
    date_filter = ""
    if start_date and end_date:
        date_filter = "WHERE r.schedule_date >= ? AND r.schedule_date < ?"
        params.extend([start_date, end_date])

    cursor.execute(f'''
        SELECT
            r.id, r.facility_type, r.schedule_date, r.start_time, r.end_time,
            r.course_code, r.course_name, r.professor, r.status,
            au.first_name, au.middle_name, au.last_name, au.program, au.year_section,
            f.code AS facility_code, p.code AS projector_code
        FROM reservation_requests r
        JOIN authorized_users au ON r.student_id = au.id
        LEFT JOIN facilities f ON r.facility_id = f.id
        LEFT JOIN projectors p ON r.projector_id = p.id
        {date_filter}
        ORDER BY r.schedule_date, r.start_time
    ''', tuple(params))
    rows = cursor.fetchall()
    conn.close()

    events = []
    for row in rows:
        title = row['facility_code'] or row['projector_code'] or row['facility_type']
        date_str = str(row['schedule_date'])
        start_time = str(row['start_time'])
        end_time = str(row['end_time'])
        if len(start_time) == 5:
            start_time = f"{start_time}:00"
        if len(end_time) == 5:
            end_time = f"{end_time}:00"
        colors = STATUS_COLORS.get(row['status'], {'backgroundColor': '#475569', 'borderColor': '#334155', 'textColor': '#ffffff'})

        middle_name = row['middle_name'] or ''
        if middle_name.strip():
            student_name = f"{row['last_name']}, {row['first_name']} {middle_name.strip()[0].upper()}."
        else:
            student_name = f"{row['last_name']}, {row['first_name']}"

        course = f"{row['course_code']} - {row['course_name']}" if row['course_code'] else 'N/A'

        events.append({
            'id': row['id'],
            'title': title,
            'start': f"{date_str}T{start_time}",
            'end': f"{date_str}T{end_time}",
            'backgroundColor': colors['backgroundColor'],
            'borderColor': colors['borderColor'],
            'textColor': colors['textColor'],
            'extendedProps': {
                'student': student_name,
                'course': course,
                'professor': row['professor'] or 'N/A',
                'status': row['status'],
                'program': row['program'] or '',
                'year_section': row['year_section'] or ''
            }
        })
    return jsonify(events)

from reportlab.pdfgen import canvas
from reportlab.lib import colors

class NumberedCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self._saved_page_states = []

    def showPage(self):
        self._saved_page_states.append(dict(self.__dict__))
        self._startPage()

    def save(self):
        num_pages = len(self._saved_page_states)
        for state in self._saved_page_states:
            self.__dict__.update(state)
            self.draw_page_number(num_pages)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)

    def draw_page_number(self, page_count):
        self.saveState()
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.HexColor("#475569"))
        self.setStrokeColor(colors.HexColor("#800000"))
        self.setLineWidth(2)
        self.line(36, 570, 756, 570)
        self.drawString(36, 575, "PUP Parañaque Resource Reservation System - Audit Trail Report")
        self.setStrokeColor(colors.HexColor("#CBD5E1"))
        self.setLineWidth(0.5)
        self.line(36, 40, 756, 40)
        page_text = f"Page {self._pageNumber} of {page_count}"
        self.drawRightString(756, 25, page_text)
        self.drawString(36, 25, "Confidential - For Administrative Use Only")
        self.restoreState()

@admin_bp.route('/admin/reports', methods=['GET', 'POST'])
def reports():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        start_date_str = request.form.get('start_date')
        end_date_str = request.form.get('end_date')
        facility_type = request.form.get('facility_type')
        format_type = request.form.get('format', 'excel')
        
        # Filter RESERVATIONS_DB
        filtered = []
        for r in RESERVATIONS_DB:
            robj = ReservationObj(r)
            if start_date_str:
                sdate = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                if robj.schedule_date < sdate:
                    continue
            if end_date_str:
                edate = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                if robj.schedule_date > edate:
                    continue
            if facility_type and robj.facility_type != facility_type:
                continue
            filtered.append(robj)
            
        if format_type == 'excel':
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            from io import BytesIO
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Utilization Report"
            
            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='800000', end_color='800000', fill_type='solid') # Maroon
            align_center = Alignment(horizontal='center', vertical='center')
            
            headers = ['ID', 'Student Name', 'Student ID', 'Program', 'Resource Category', 'Resource Code', 'Date', 'Start Time', 'End Time', 'Course & Professor', 'Status', 'Remarks']
            ws.append(headers)
            
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                
            for r in filtered:
                code = r.facility.code if r.facility else (r.projector.code if r.projector else 'N/A')
                student_name = r.user.full_name if r.user else 'Guest'
                student_num = r.user.student_number if r.user else 'N/A'
                prog = r.user.program if r.user else 'N/A'
                course_prof = f"{r.course_code or 'N/A'} (Prof. {r.professor or 'N/A'})"
                
                row_data = [
                    r.id,
                    student_name,
                    student_num,
                    prog,
                    r.facility_type,
                    code,
                    r.schedule_date.strftime('%Y-%m-%d'),
                    r.start_time.strftime('%I:%M %p'),
                    r.end_time.strftime('%I:%M %p'),
                    course_prof,
                    r.status,
                    r.remarks or ''
                ]
                ws.append(row_data)
                
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
            out = BytesIO()
            wb.save(out)
            out.seek(0)
            
            response = Response(out.getvalue(), mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response.headers['Content-Disposition'] = 'attachment; filename=PUPPQ_ReservationSystem_Report.xlsx'
            return response
            
        elif format_type == 'pdf':
            from reportlab.lib.pagesizes import letter, landscape
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from io import BytesIO
            
            out = BytesIO()
            doc = SimpleDocTemplate(
                out,
                pagesize=landscape(letter),
                leftMargin=36,
                rightMargin=36,
                topMargin=54,
                bottomMargin=54
            )
            
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontName='Helvetica-Bold',
                fontSize=20,
                leading=24,
                textColor=colors.HexColor("#800000"),
                spaceAfter=6
            )
            
            meta_style = ParagraphStyle(
                'ReportMeta',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=10,
                leading=14,
                textColor=colors.HexColor("#475569")
            )
            
            cell_style = ParagraphStyle(
                'TableCell',
                parent=styles['Normal'],
                fontName='Helvetica',
                fontSize=8,
                leading=10,
                textColor=colors.HexColor("#1E293B")
            )
            cell_style_bold = ParagraphStyle(
                'TableCellBold',
                parent=cell_style,
                fontName='Helvetica-Bold'
            )
            cell_header_style = ParagraphStyle(
                'TableHeader',
                parent=cell_style,
                fontName='Helvetica-Bold',
                textColor=colors.white,
                alignment=1
            )
            
            story = []
            
            # Title
            story.append(Paragraph("RESOURCE UTILIZATION AUDIT REPORT", title_style))
            
            # Metadata
            filters_str = []
            if start_date_str:
                filters_str.append(f"Start Date: {start_date_str}")
            if end_date_str:
                filters_str.append(f"End Date: {end_date_str}")
            if facility_type:
                filters_str.append(f"Resource Category: {facility_type}")
            else:
                filters_str.append("Resource Category: All")
                
            filters_txt = ", ".join(filters_str)
            meta_txt = f"<b>Generated On:</b> {datetime.now().strftime('%B %d, %Y %I:%M %p')}<br/><b>Filters Applied:</b> {filters_txt}<br/><b>Total Records Matched:</b> {len(filtered)}"
            
            story.append(Paragraph(meta_txt, meta_style))
            story.append(Spacer(1, 15))
            
            # Table Data
            table_data = []
            headers = ['ID', 'Student Name', 'Student ID', 'Category', 'Code', 'Date', 'Times', 'Course & Professor', 'Status', 'Remarks']
            table_data.append([Paragraph(h, cell_header_style) for h in headers])
            
            for r in filtered:
                code = r.facility.code if r.facility else (r.projector.code if r.projector else 'N/A')
                student_name = r.user.full_name if r.user else 'Guest'
                student_num = r.user.student_number if r.user else 'N/A'
                prog = r.user.program if r.user else 'N/A'
                course_prof = f"{r.course_code or 'N/A'}<br/><font color='#64748B'>Prof. {r.professor or 'N/A'}</font>"
                times_str = f"{r.start_time.strftime('%I:%M %p')} - {r.end_time.strftime('%I:%M %p')}"
                
                status_color = '#10B981'
                if r.status == 'PENDING APPROVAL':
                    status_color = '#F59E0B'
                elif r.status == 'REJECTED':
                    status_color = '#EF4444'
                elif r.status == 'AWAITING RETURN':
                    status_color = '#7C3AED'
                elif r.status == 'OVERDUE':
                    status_color = '#B91C1C'
                elif r.status == 'CANCELLED':
                    status_color = '#6B7280'
                elif r.status == 'COMPLETED':
                    status_color = "#0284C7"
                    
                status_html = f"<b><font color='{status_color}'>{r.status}</font></b>"
                
                row = [
                    Paragraph(f"#{r.id}", cell_style_bold),
                    Paragraph(f"<b>{student_name}</b><br/><font size='7' color='#64748B'>{prog}</font>", cell_style),
                    Paragraph(student_num, cell_style),
                    Paragraph(r.facility_type, cell_style),
                    Paragraph(code, cell_style_bold),
                    Paragraph(r.schedule_date.strftime('%Y-%m-%d'), cell_style),
                    Paragraph(times_str, cell_style),
                    Paragraph(course_prof, cell_style),
                    Paragraph(status_html, cell_style),
                    Paragraph(r.remarks or '', cell_style)
                ]
                table_data.append(row)
                
            col_widths = [30, 105, 60, 80, 50, 65, 95, 120, 65, 50]
            t = Table(table_data, colWidths=col_widths, repeatRows=1)
            
            t_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#800000")),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ])
            
            for i in range(1, len(table_data)):
                bg_color = colors.HexColor("#F8FAFC") if i % 2 == 1 else colors.white
                t_style.add('BACKGROUND', (0, i), (-1, i), bg_color)
                
            t.setStyle(t_style)
            story.append(t)
            
            doc.build(story, canvasmaker=NumberedCanvas)
            out.seek(0)
            
            response = Response(out.getvalue(), mimetype='application/pdf')
            response.headers['Content-Disposition'] = 'attachment; filename=PUPPQ_ReservationSystem_Report.pdf'
            return response
            
        else:
            # Export as CSV with formatted header metadata
            si = StringIO()
            cw = csv.writer(si)
            # === Metadata Header Block ===
            cw.writerow(['PUP Parañaque Resource Reservation System'])
            cw.writerow(['Resource Utilization Audit Report'])
            cw.writerow([])
            cw.writerow(['Generated On:', datetime.now().strftime('%B %d, %Y %I:%M %p')])
            # Filters
            filter_parts = []
            if start_date_str:
                filter_parts.append(f'Start Date: {start_date_str}')
            if end_date_str:
                filter_parts.append(f'End Date: {end_date_str}')
            if facility_type:
                filter_parts.append(f'Resource Category: {facility_type}')
            else:
                filter_parts.append('Resource Category: All')
            cw.writerow(['Filters Applied:', ', '.join(filter_parts)])
            cw.writerow(['Total Records Matched:', len(filtered)])
            cw.writerow([])
            # === Data Grid ===
            cw.writerow(['ID', 'Student Name', 'Student ID', 'Program', 'Resource Category', 'Resource Code', 'Date', 'Start Time', 'End Time', 'Course & Professor', 'Status', 'Remarks'])
            
            for r in filtered:
                code = r.facility.code if r.facility else (r.projector.code if r.projector else 'N/A')
                student_name = r.user.full_name if r.user else 'Guest'
                student_num = r.user.student_number if r.user else 'N/A'
                prog = r.user.program if r.user else 'N/A'
                course_prof = f"{r.course_code or 'N/A'} (Prof. {r.professor or 'N/A'})"
                
                cw.writerow([
                    r.id,
                    student_name,
                    student_num,
                    prog,
                    r.facility_type,
                    code,
                    r.schedule_date.strftime('%Y-%m-%d'),
                    r.start_time.strftime('%I:%M %p'),
                    r.end_time.strftime('%I:%M %p'),
                    course_prof,
                    r.status,
                    r.remarks or ''
                ])
                
            response = Response(si.getvalue(), mimetype='text/csv')
            response.headers['Content-Disposition'] = 'attachment; filename=PUPPQ_ReservationSystem_Report.csv'
            return response
            
    return render_template('admin/reports.html', form=MockForm())

@admin_bp.route('/admin/ai-insights', methods=['GET', 'POST'])
def ai_insights():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))

    if request.method == 'POST':
        buffer_minutes = request.form.get('buffer_minutes', DEFAULT_AI_SETTINGS['buffer_minutes'])
        auto_suggest = request.form.get('auto_suggest') == '1'
        peak_warning = request.form.get('peak_warning') == '1'
        email_alerts = request.form.get('email_alerts') == '1'
        try:
            save_ai_settings(buffer_minutes, auto_suggest, peak_warning, email_alerts)
            log_event(u.id, "UPDATE AI SETTINGS", "Updated AI recommendation configuration.")
            flash("AI configuration saved successfully.", "success")
        except Exception as e:
            flash(f"Unable to save AI configuration: {e}", "danger")
        return redirect(url_for('admin.ai_insights'))

    ai_settings = load_ai_settings()

    def parse_date(value):
        if hasattr(value, 'weekday'):
            return value
        return datetime.strptime(str(value), '%Y-%m-%d').date()

    def parse_time(value):
        if hasattr(value, 'hour'):
            return value
        text = str(value)
        return datetime.strptime(text, '%H:%M:%S' if text.count(':') == 2 else '%H:%M').time()

    def fmt_hour(hour):
        return f"{hour % 12 or 12}:00 {'AM' if hour < 12 else 'PM'}"

    def category_name(raw_type):
        return 'Projector' if raw_type == 'Projector (Equipment Only)' else raw_type

    def risk_level(percent):
        if percent >= 80:
            return 'High'
        if percent >= 55:
            return 'Medium'
        return 'Low'

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute('''
        SELECT facility_type, schedule_date, start_time, end_time, status
        FROM reservation_requests
        WHERE status IN ('APPROVED', 'PENDING APPROVAL', 'IN USE')
    ''')
    reservations = []
    for row in cursor.fetchall():
        schedule_date = parse_date(row['schedule_date'])
        start_time = parse_time(row['start_time'])
        end_time = parse_time(row['end_time'])
        start_hour = start_time.hour
        end_hour = max(start_hour + 1, end_time.hour + (1 if end_time.minute else 0))
        reservations.append({
            'category': category_name(row['facility_type']),
            'date': schedule_date,
            'weekday': schedule_date.weekday(),
            'start_hour': start_hour,
            'end_hour': min(end_hour, 22),
            'status': row['status']
        })

    cursor.execute("SELECT type, COUNT(*) AS total FROM facilities WHERE status != 'Archived' GROUP BY type")
    category_capacity = defaultdict(int)
    for row in cursor.fetchall():
        category_capacity[category_name(row['type'])] += row['total']
    cursor.execute("SELECT COUNT(*) FROM projectors WHERE status != 'Archived'")
    category_capacity['Projector'] += cursor.fetchone()[0]
    conn.close()

    categories = ['Audio-Visual Room (AVR)', 'Computer Laboratory', 'Hospitality Management Laboratory', 'Projector']
    today = datetime.now().date()
    weekday_names = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']

    day_counts = Counter()
    hour_counts = Counter()
    category_counts = Counter()
    category_hours = Counter()
    category_peak_slot = {}
    category_hour_counts = defaultdict(Counter)
    recent_30 = 0
    previous_30 = 0

    for res in reservations:
        day_counts[res['weekday']] += 1
        category_counts[res['category']] += 1
        days_old = (today - res['date']).days
        if 0 <= days_old < 30:
            recent_30 += 1
        elif 30 <= days_old < 60:
            previous_30 += 1

        for hour in range(max(7, res['start_hour']), min(22, res['end_hour'])):
            hour_counts[hour] += 1
            category_hour_counts[res['category']][hour] += 1
            category_hours[res['category']] += 1

    peak_day_idx = day_counts.most_common(1)[0][0] if day_counts else today.weekday()
    peak_day = weekday_names[peak_day_idx]
    peak_hour = hour_counts.most_common(1)[0][0] if hour_counts else 10
    peak_hour_val = f"{fmt_hour(peak_hour)} - {fmt_hour(peak_hour + 1)}"

    trend_factor = 1.0
    if previous_30 > 0:
        trend_factor = max(0.5, min(1.75, recent_30 / previous_30))
    elif recent_30 > 0:
        trend_factor = 1.2

    category_index = {category: idx for idx, category in enumerate(categories)}
    daily_counts = Counter((res['date'], res['category']) for res in reservations)
    ml_enabled = False
    ml_note = "Machine learning model is waiting for reservation history before producing forecasts."
    ml_confidence = 0
    ml_predictions = {category: {'max': 0, 'day': peak_day, 'hour': peak_hour, 'avg': 0} for category in categories}

    if daily_counts:
        try:
            from sklearn.linear_model import LinearRegression
            from sklearn.metrics import mean_absolute_error

            x_train = []
            y_train = []
            for (sample_date, category), count in daily_counts.items():
                x_train.append([
                    sample_date.toordinal(),
                    sample_date.weekday(),
                    sample_date.month,
                    1 if sample_date.weekday() >= 5 else 0,
                    category_index.get(category, len(categories))
                ])
                y_train.append(count)

            model = LinearRegression()
            model.fit(x_train, y_train)
            fitted = model.predict(x_train)
            mae = mean_absolute_error(y_train, fitted) if len(y_train) > 1 else 0
            avg_target = max(1, sum(y_train) / len(y_train))
            ml_confidence = max(45, min(96, round(100 - (mae / avg_target * 35))))
            ml_enabled = True
            ml_note = "Linear Regression model trained on historical daily reservation demand."

            for category in categories:
                future_rows = []
                for offset in range(1, 15):
                    forecast_date = today + timedelta(days=offset)
                    future_rows.append([
                        forecast_date.toordinal(),
                        forecast_date.weekday(),
                        forecast_date.month,
                        1 if forecast_date.weekday() >= 5 else 0,
                        category_index.get(category, len(categories))
                    ])

                predicted_values = [max(0, float(v)) for v in model.predict(future_rows)]
                max_idx = max(range(len(predicted_values)), key=lambda idx: predicted_values[idx])
                forecast_date = today + timedelta(days=max_idx + 1)
                ml_predictions[category] = {
                    'max': predicted_values[max_idx],
                    'avg': sum(predicted_values) / len(predicted_values),
                    'day': weekday_names[forecast_date.weekday()],
                    'hour': category_hour_counts[category].most_common(1)[0][0] if category_hour_counts[category] else peak_hour
                }
        except Exception as e:
            ml_note = f"Machine learning forecast unavailable: {e}"

    forecast_rows = []
    highest_forecast = 0
    total_capacity_hours = 0
    total_booked_hours = 0
    unique_days = max(1, len({r['date'] for r in reservations}))

    for category in categories:
        capacity_units = max(1, category_capacity.get(category, 0))
        capacity_hours = capacity_units * 14 * unique_days
        booked_hours = category_hours.get(category, 0)
        total_capacity_hours += capacity_hours
        total_booked_hours += booked_hours
        avg_occupancy = round((booked_hours / capacity_hours) * 100) if capacity_hours else 0

        peak_slot_hour = category_hour_counts[category].most_common(1)[0][0] if category_hour_counts[category] else peak_hour
        category_peak_slot[category] = peak_slot_hour

        base_daily_demand = category_counts.get(category, 0) / unique_days
        ml_daily_demand = ml_predictions[category]['max'] if ml_enabled else base_daily_demand * trend_factor
        forecast_pct = min(100, round((ml_daily_demand / capacity_units) * 100))
        highest_forecast = max(highest_forecast, forecast_pct)
        level = risk_level(max(avg_occupancy, forecast_pct))

        forecast_rows.append({
            'category': category,
            'avg_occupancy': f"{avg_occupancy}%",
            'predicted_demand': f"{ml_daily_demand:.1f} bookings/day",
            'peak_forecast': f"{ml_predictions[category]['day'] if ml_enabled else peak_day} {fmt_hour(ml_predictions[category]['hour'] if ml_enabled else peak_slot_hour)}",
            'risk_level': level,
            'risk_class': 'danger' if level == 'High' else ('warning text-dark' if level == 'Medium' else 'success')
        })

    lowest_day_idx = min(range(7), key=lambda idx: day_counts.get(idx, 0))
    lowest_hour = min(range(7, 22), key=lambda hour: hour_counts.get(hour, 0))
    maintenance_window = f"{weekday_names[lowest_day_idx]} {fmt_hour(lowest_hour)}"

    upcoming_72_cutoff = today + timedelta(days=3)
    upcoming_count = sum(1 for r in reservations if today <= r['date'] <= upcoming_72_cutoff)
    alert_level = 'High' if upcoming_count >= 12 or highest_forecast >= 80 else ('Watch' if upcoming_count >= 6 or highest_forecast >= 55 else 'Safe')
    alert_class = 'danger' if alert_level == 'High' else ('warning text-dark' if alert_level == 'Watch' else 'success')
    alert_text = (
        f"{upcoming_count} active bookings are scheduled in the next 72 hours. "
        f"Forecast pressure is {highest_forecast}% based on recent demand and available resources."
    )

    recommendation_items = []
    if alert_level == 'High':
        recommendation_items.append("Temporarily steer new requests away from the peak window and require manual review for overlapping resource categories.")
    elif alert_level == 'Watch':
        recommendation_items.append("Show peak-utilization warnings during the busiest hour and prioritize alternative slot suggestions.")
    else:
        recommendation_items.append("Keep automatic recommendations enabled; current booking pressure is within normal operating range.")

    busiest_category = category_counts.most_common(1)[0][0] if category_counts else 'Projector'
    recommendation_items.append(f"Pre-stage support for {busiest_category} around {peak_hour_val}.")
    recommendation_items.append(f"Schedule low-impact maintenance around {maintenance_window}.")

    model_confidence = ml_confidence if ml_enabled else min(70, 45 + len(reservations) * 3)
    recommended_buffer_minutes = 30 if alert_level == 'High' else (20 if alert_level == 'Watch' else 15)

    insights = {
        'peak_day': peak_day,
        'peak_day_reason': f"{peak_day} has the strongest historical booking signal in the current dataset.",
        'peak_hour': peak_hour_val,
        'peak_hour_reason': f"Bookings cluster most often around {peak_hour_val}.",
        'maintenance_window': maintenance_window,
        'maintenance_reason': "This is the lowest-demand day and hour combination found from reservation history.",
        'alert_level': alert_level,
        'alert_class': alert_class,
        'alert_text': alert_text,
        'forecast_rows': forecast_rows,
        'recommendations': recommendation_items,
        'model_confidence': model_confidence,
        'model_type': 'Linear Regression' if ml_enabled else 'Heuristic fallback',
        'model_note': ml_note,
        'ml_enabled': ml_enabled,
        'buffer_minutes': ai_settings['buffer_minutes'],
        'recommended_buffer_minutes': recommended_buffer_minutes,
        'settings': ai_settings,
        'total_samples': len(reservations),
        'trend_label': 'Rising' if trend_factor > 1.1 else ('Cooling' if trend_factor < 0.9 else 'Stable'),
        'generated_at': datetime.now().strftime('%B %d, %Y %I:%M %p')
    }
    return render_template('admin/ai_insights.html', insights=insights)

@admin_bp.route('/admin/logs')
def logs():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    logs_objs = [SystemLogObj(l) for l in sorted(system_logs, key=lambda x: x['created_at'], reverse=True)]
    return render_template('admin/logs.html', logs=logs_objs)

@admin_bp.route('/admin/settings', methods=['GET', 'POST'])
def settings():
    u = get_current_user()
    if not u.is_authenticated or u.role != 'admin':
        return redirect(url_for('auth.login'))
        
    if request.method == 'POST':
        current_password = request.form.get('current_password', '').strip()
        new_password = request.form.get('new_password', '').strip()
        confirm_new_password = request.form.get('confirm_new_password', '').strip()
        
        user_dict = None
        for usr in USERS_DB:
            if usr['id'] == u.id:
                user_dict = usr
                break
                
        if user_dict and new_password:
            if not check_password_hash(user_dict['password_hash'], current_password):
                return render_template('admin/settings.html', form=MockForm(), current_password_error="Current password is incorrect.")

            if check_password_hash(user_dict['password_hash'], new_password):
                return render_template('admin/settings.html', form=MockForm(), new_password_error="New password cannot be the same as your current password.")

            if new_password != confirm_new_password:
                return render_template('admin/settings.html', form=MockForm(), confirm_password_error="Passwords should match.")
                
            new_hash = generate_password_hash(new_password)
            
            # Persist back to admin table in database
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute("UPDATE admin SET password_hash = ? WHERE username = ?", (new_hash, u.id))
            conn.commit()
            conn.close()
            
            # Sync memory cache
            sync_db_to_memory()
            
            log_event(u.id, "UPDATE SETTINGS", "Administrator changed account password.")
            flash("Settings saved successfully!", "success")
            
    return render_template('admin/settings.html', form=MockForm())


# ==========================================
# AI Chatbot Assistant Route
# ==========================================
FACILITY_ALIASES = {
    'avr': 'Audio-Visual Room (AVR)',
    'audio visual': 'Audio-Visual Room (AVR)',
    'audio-visual': 'Audio-Visual Room (AVR)',
    'av room': 'Audio-Visual Room (AVR)',
    'computer': 'Computer Laboratory',
    'comp': 'Computer Laboratory',
    'comp lab': 'Computer Laboratory',
    'computer lab': 'Computer Laboratory',
    'com lab': 'Computer Laboratory',
    'laboratory': 'Computer Laboratory',
    'hml': 'Hospitality Management Laboratory',
    'hm': 'Hospitality Management Laboratory',
    'hm lab': 'Hospitality Management Laboratory',
    'hospitality lab': 'Hospitality Management Laboratory',
    'hospitality': 'Hospitality Management Laboratory',
    'projector': 'Projector',
    'projectors': 'Projector'
}

CHATBOT_SUGGESTIONS = {
    'guest': ['Reservation requirements', 'Available rooms tomorrow', 'How to reserve'],
    'student': ['My reservation status', 'Available AVR tomorrow 10 AM', 'Reservation requirements'],
    'admin': ['Today summary', 'Pending approvals', 'Most requested facility']
}

def _clean_chat_html(text):
    if not text:
        return ''
    for src, repl in {'<br>': '\n', '<br/>': '\n', '<br />': '\n', '<strong>': '', '</strong>': '', '<b>': '', '</b>': ''}.items():
        text = text.replace(src, repl)
    return re.sub(r'<[^>]+>', '', text).strip()

def _chat_role(user):
    if not user or not user.is_authenticated:
        return 'guest'
    return 'admin' if user.role == 'admin' else 'student'

def _normalized_chat_text(message):
    return re.sub(r'[^a-z0-9\s]', '', message.lower()).strip()

def _detect_chat_intent(message):
    msg = message.lower()
    simple_msg = _normalized_chat_text(message)
    greeting_words = {'hi', 'hello', 'hey', 'good morning', 'good afternoon', 'good evening', 'mabuhay'}
    ack_words = {'ok', 'okay', 'oki', 'alright', 'got it', 'noted', 'sure', 'yes', 'yep'}
    thanks_words = {'thanks', 'thank you', 'ty', 'salamat'}
    bye_words = {'bye', 'goodbye', 'see you'}
    if simple_msg in greeting_words:
        return 'greeting'
    if simple_msg in ack_words:
        return 'acknowledgement'
    if simple_msg in thanks_words:
        return 'thanks'
    if simple_msg in bye_words:
        return 'goodbye'
    if _detect_facility_type(message):
        if any(word in msg for word in ('available', 'availability', 'free', 'vacant', 'slot', 'schedule')):
            return 'availability'
        return 'room_info'
    if any(word in msg for word in ('available', 'availability', 'free', 'vacant', 'slot', 'schedule')):
        return 'availability'
    if any(word in msg for word in ('requirement', 'requirements', 'needed', 'need to submit', 'documents')):
        return 'requirements'
    if any(word in msg for word in ('reject', 'rejection', 'reason', 'remarks')):
        return 'rejection_reason'
    if any(word in msg for word in ('summary', 'summarize', 'insight', 'activity', 'most requested', 'pending approval', 'pending approvals')):
        return 'admin_insights'
    if any(word in msg for word in ('my reservation', 'my booking', 'status', 'pending', 'approved', 'rejected')):
        return 'status'
    if any(word in msg for word in ('how', 'guide', 'reserve', 'book', 'submit', 'request')):
        return 'guide'
    if any(word in msg for word in ('cancel', 'modify', 'change', 'edit')):
        return 'manage'
    if any(word in msg for word in ('capacity', 'equipment', 'room')):
        return 'room_info'
    return 'general'

def _detect_facility_type(message):
    msg = _normalized_chat_text(message)
    for alias, facility_type in sorted(FACILITY_ALIASES.items(), key=lambda item: len(item[0]), reverse=True):
        normalized_alias = _normalized_chat_text(alias)
        if re.search(rf'(^|\s){re.escape(normalized_alias)}(\s|$)', msg):
            return facility_type
    return None

def _parse_chat_date(message):
    msg = message.lower()
    today = datetime.now().date()
    if 'tomorrow' in msg:
        return today + timedelta(days=1)
    if 'today' in msg:
        return today
    match = re.search(r'(\d{4}-\d{2}-\d{2})', msg)
    if match:
        return datetime.strptime(match.group(1), '%Y-%m-%d').date()
    match = re.search(r'\b(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\.?\s+(\d{1,2})(?:,\s*(\d{4}))?', msg)
    if match:
        month_lookup = {
            'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
            'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12
        }
        return date(int(match.group(3) or today.year), month_lookup[match.group(1)[:3]], int(match.group(2)))
    return None

def _parse_chat_times(message):
    msg = message.lower().replace('.', '')
    matches = list(re.finditer(r'(\d{1,2})(?::(\d{2}))?\s*(am|pm)?', msg))
    parsed = []
    for match in matches:
        hour = int(match.group(1))
        minute = int(match.group(2) or 0)
        meridiem = match.group(3)
        if hour > 24 or minute > 59:
            continue
        if meridiem == 'pm' and hour != 12:
            hour += 12
        elif meridiem == 'am' and hour == 12:
            hour = 0
        if 0 <= hour <= 23:
            parsed.append(time(hour, minute))
    if len(parsed) >= 2:
        return parsed[0], parsed[1]
    if len(parsed) == 1:
        return parsed[0], (datetime.combine(datetime.now().date(), parsed[0]) + timedelta(hours=1)).time()
    return None, None

def _format_chat_time(value):
    return value.strftime('%I:%M %p').lstrip('0')

def _facility_capacity(facility_type):
    return {
        'Audio-Visual Room (AVR)': 'approximately 80-100 people',
        'Computer Laboratory': 'approximately 30-40 people',
        'Hospitality Management Laboratory': 'approximately 25-35 people',
        'Projector': 'portable equipment; capacity depends on the assigned room'
    }.get(facility_type, 'capacity details are not configured')

def _facility_equipment(facility_type):
    return {
        'Audio-Visual Room (AVR)': 'AV equipment, audio controls, display/projector support, and air-conditioning',
        'Computer Laboratory': 'desktop computers and laboratory teaching equipment',
        'Hospitality Management Laboratory': 'hospitality practical training equipment',
        'Projector': 'portable projector units from the projector fleet'
    }.get(facility_type, 'equipment details are not configured')

def _room_is_available(room, target_date, start_time, end_time):
    if room['status'] in FACILITY_HARD_BLOCK_STATUSES:
        return False
    for res in RESERVATIONS_DB:
        if res['status'] in ('APPROVED', 'PENDING APPROVAL', 'IN USE') and res.get('facility_id') == room['id']:
            if res['schedule_date'] == target_date and start_time < res['end_time'] and end_time > res['start_time']:
                return False
    return True

def _projector_is_available(projector, target_date, start_time, end_time):
    if projector['status'] in PROJECTOR_HARD_BLOCK_STATUSES:
        return False
    for res in RESERVATIONS_DB:
        if res['status'] in ('APPROVED', 'PENDING APPROVAL') and res.get('projector_id') == projector['id']:
            if res['schedule_date'] == target_date and start_time < res['end_time'] and end_time > res['start_time']:
                return False
    return True

def _chat_availability_context(message):
    facility_type = _detect_facility_type(message)
    target_date = _parse_chat_date(message)
    start_time, end_time = _parse_chat_times(message)
    context = {
        'intent': 'availability',
        'facility_type': facility_type,
        'date': target_date.strftime('%Y-%m-%d') if target_date else None,
        'start_time': start_time.strftime('%H:%M') if start_time else None,
        'end_time': end_time.strftime('%H:%M') if end_time else None,
        'operating_hours': '7:00 AM to 9:00 PM',
        'blocking_statuses': ['APPROVED', 'PENDING APPROVAL', 'IN USE'],
        'available': [],
        'unavailable': [],
        'resource_total': 0,
        'resource_available_now': 0,
        'needs_more_information': []
    }
    if not target_date:
        context['needs_more_information'].append('date')
    if not start_time or not end_time:
        context['needs_more_information'].append('time range')
    if not facility_type:
        context['needs_more_information'].append('facility type')
    if facility_type == 'Projector':
        context['resource_total'] = len([p for p in PROJECTORS_DB if p['status'] != 'Archived'])
        context['resource_available_now'] = len([p for p in PROJECTORS_DB if p['status'] == 'Available'])
    elif facility_type:
        rooms_for_type = [f for f in FACILITIES_DB if f['type'] == facility_type and f['status'] not in FACILITY_HARD_BLOCK_STATUSES]
        context['resource_total'] = len(rooms_for_type)
        context['resource_available_now'] = len([f for f in rooms_for_type if f['status'] == 'Available'])
    if context['needs_more_information']:
        return context
    if start_time >= end_time or start_time < time(7, 0) or end_time > time(21, 0):
        context['outside_operating_hours'] = True
        return context

    if facility_type == 'Projector':
        for projector in PROJECTORS_DB:
            row = {'code': projector['code'], 'status': projector['status'], 'equipment': projector['model']}
            if _projector_is_available(projector, target_date, start_time, end_time):
                context['available'].append(row)
            else:
                context['unavailable'].append(row)
    else:
        for room in [f for f in FACILITIES_DB if f['type'] == facility_type and f['status'] != 'Archived']:
            row = {
                'code': room['code'],
                'status': room['status'],
                'capacity': _facility_capacity(facility_type),
                'equipment': _facility_equipment(facility_type)
            }
            if _room_is_available(room, target_date, start_time, end_time):
                context['available'].append(row)
            else:
                context['unavailable'].append(row)
    return context

def _chat_status_context(user, intent):
    context = {'intent': intent, 'reservations': []}
    if not user.is_authenticated or user.role != 'student':
        context['needs_login'] = True
        return context
    rows = [r for r in RESERVATIONS_DB if r.get('user_id') == user.id]
    rows.sort(key=lambda r: (r['schedule_date'], r['start_time']), reverse=True)
    for res in rows[:6]:
        resource = 'N/A'
        if res.get('facility_id'):
            resource = next((f['code'] for f in FACILITIES_DB if f['id'] == res['facility_id']), res['facility_type'])
        elif res.get('projector_id'):
            resource = next((p['code'] for p in PROJECTORS_DB if p['id'] == res['projector_id']), res['facility_type'])
        context['reservations'].append({
            'id': res['id'],
            'facility_type': res['facility_type'],
            'resource': resource,
            'date': res['schedule_date'].strftime('%Y-%m-%d'),
            'time': f"{_format_chat_time(res['start_time'])} - {_format_chat_time(res['end_time'])}",
            'status': res['status'],
            'remarks': res.get('remarks') or ''
        })
    return context

def _chat_admin_context():
    today = datetime.now().date()
    month_start = today.replace(day=1)
    today_rows = [r for r in RESERVATIONS_DB if r['schedule_date'] == today]
    month_rows = [r for r in RESERVATIONS_DB if r['schedule_date'] >= month_start]
    requested = Counter(r['facility_type'] for r in month_rows)
    return {
        'intent': 'admin_insights',
        'today': today.strftime('%Y-%m-%d'),
        'today_total': len(today_rows),
        'today_pending': sum(1 for r in today_rows if r['status'] == 'PENDING APPROVAL'),
        'all_pending': sum(1 for r in RESERVATIONS_DB if r['status'] == 'PENDING APPROVAL'),
        'all_approved': sum(1 for r in RESERVATIONS_DB if r['status'] == 'APPROVED'),
        'active_booking_count': sum(1 for r in RESERVATIONS_DB if r['status'] in ('APPROVED', 'PENDING APPROVAL', 'IN USE')),
        'most_requested_facility_this_month': requested.most_common(1)[0][0] if requested else 'No reservation requests yet'
    }

def _chat_room_context(message):
    facility_type = _detect_facility_type(message)
    facilities = []
    if facility_type == 'Projector':
        facilities = [{'code': p['code'], 'status': p['status'], 'equipment': p['model']} for p in PROJECTORS_DB]
    else:
        for room in [f for f in FACILITIES_DB if not facility_type or f['type'] == facility_type]:
            facilities.append({
                'code': room['code'],
                'type': room['type'],
                'status': room['status'],
                'capacity': _facility_capacity(room['type']),
                'equipment': _facility_equipment(room['type'])
            })
    return {'intent': 'room_info', 'facility_type': facility_type, 'facilities': facilities[:12]}

def _chat_guide_context(intent):
    return {
        'intent': intent,
        'supported_facilities': [
            'Audio-Visual Room (AVR)',
            'Computer Laboratory',
            'Hospitality Management Laboratory',
            'Projector'
        ],
        'reservation_rules': [
            'Students must log in before submitting a request.',
            'Choose facility type, date, start time, end time, resource, course code, course name, and professor.',
            'Reservations must be between 7:00 AM and 9:00 PM.',
            'Approved and pending reservations block overlapping slots.',
            'Requests are submitted as PENDING APPROVAL until an admin approves or rejects them.'
        ],
        'requirements': [
            'Active student account in the reservation system.',
            'Facility or projector category.',
            'Schedule date plus start and end time.',
            'Available room or projector unit selected from the system.',
            'Course code, course name, and professor or supervising faculty.',
            'Clear academic purpose for the reservation when requested by admin.'
        ],
        'student_next_action': 'Open New Reservation and complete the booking form.',
        'admin_next_action': 'Open Requests Queue to review pending approvals.'
    }

def _build_chat_context(user, message, intent):
    role = _chat_role(user)
    context = {
        'app': 'PUP Paranaque Reservation System',
        'user_role': role,
        'current_date': datetime.now().strftime('%Y-%m-%d'),
        'intent': intent
    }
    if intent == 'availability':
        context.update(_chat_availability_context(message))
    elif intent in ('status', 'rejection_reason'):
        context.update(_chat_status_context(user, intent))
    elif intent == 'admin_insights':
        if role == 'admin':
            context.update(_chat_admin_context())
        else:
            context['not_allowed'] = 'Admin insights are only available to administrators.'
    elif intent == 'room_info':
        context.update(_chat_room_context(message))
    elif intent in ('guide', 'manage', 'requirements'):
        context.update(_chat_guide_context(intent))
    elif intent in ('greeting', 'acknowledgement', 'thanks', 'goodbye'):
        context['conversation_turn'] = intent
    else:
        context['fallback_suggestions'] = CHATBOT_SUGGESTIONS.get(role, CHATBOT_SUGGESTIONS['guest'])
    return context

def _basic_chatbot_response(context):
    intent = context.get('intent')
    role = context.get('user_role', 'guest')
    if intent == 'greeting':
        if role == 'admin':
            return "Hello! I’m ready to help with today’s reservation activity, pending approvals, facility demand, or schedule checks. You can try: <strong>Today summary</strong> or <strong>Pending approvals</strong>."
        if role == 'student':
            return "Hi! I can help you check room availability, view your reservation status, or guide you through a new booking. You can ask: <strong>Is AVR available tomorrow at 10 AM?</strong>"
        return "Hi! I’m the PUP Reservation Assistant. I can explain requirements, help with facility availability, and guide you through the reservation process."
    if intent == 'acknowledgement':
        return "Got it. I’m here if you want to check availability, review a reservation status, or ask about requirements."
    if intent == 'thanks':
        return "You’re welcome. I can also help with available time slots, room capacity, equipment, or booking steps."
    if intent == 'goodbye':
        return "Goodbye! Come back anytime you need help with reservations or schedules."
    if intent == 'availability':
        if context.get('outside_operating_hours'):
            return "That time is outside reservation hours. The system accepts bookings from <strong>7:00 AM to 9:00 PM</strong> only.<br><br><strong>Try instead:</strong> ask for a slot like \"How many projectors are available tomorrow at 10 AM?\" or \"Projectors tomorrow 1 PM to 3 PM.\""
        if context.get('needs_more_information'):
            missing = context['needs_more_information']
            facility = context.get('facility_type')
            date_value = context.get('date')
            if facility and missing == ['time range']:
                date_text = datetime.strptime(date_value, '%Y-%m-%d').strftime('%B %d, %Y') if date_value else 'that date'
                total = context.get('resource_total', 0)
                available_now = context.get('resource_available_now', 0)
                resource_label = 'projector(s)' if facility == 'Projector' else facility
                return f"I found the resource type: <strong>{facility}</strong> for <strong>{date_text}</strong>.<br><br>The system currently has <strong>{available_now} of {total}</strong> {resource_label} marked available, but exact availability depends on the time because reservations can overlap by hour.<br><br><strong>Please include a time range</strong>, for example: \"How many projectors are available tomorrow 10 AM to 12 PM?\""
            needed = ', '.join(missing)
            example = "How many projectors are available tomorrow 10 AM to 12 PM?" if facility == 'Projector' else "Is AVR available tomorrow 10 AM to 12 PM?"
            return f"I can check that in the database. Please include the {needed}.<br><br>Example: \"{example}\""
        available = context.get('available', [])
        facility = context.get('facility_type') or 'resource'
        date_text = datetime.strptime(context['date'], '%Y-%m-%d').strftime('%B %d, %Y')
        start = _format_chat_time(datetime.strptime(context['start_time'], '%H:%M').time())
        end = _format_chat_time(datetime.strptime(context['end_time'], '%H:%M').time())
        if available:
            codes = ', '.join(item['code'] for item in available[:5])
            details = ''
            if available[0].get('capacity') or available[0].get('equipment'):
                details = f"<br><br><strong>Room details:</strong><br>Capacity: {available[0].get('capacity', 'N/A')}<br>Equipment: {available[0].get('equipment', 'N/A')}"
            return f"Yes. I checked the live reservation records for {date_text}, {start} to {end}.<br><br><strong>Available {facility}:</strong> {codes}<br><br>Pending and approved bookings are treated as blocked, so this slot should be safe to request.{details}<br><br><strong>Next step:</strong> open New Reservation, choose the same date/time, then select {available[0]['code']}."
        return f"I checked the database and did not find an available {facility} on {date_text}, {start} to {end}.<br><br>That means every matching resource is either already approved, pending approval, or unavailable for that time.<br><br><strong>Try next:</strong> ask for another time range, such as \"available AVR tomorrow 1 PM to 3 PM,\" or choose a different facility."
    if intent in ('status', 'rejection_reason'):
        if context.get('needs_login'):
            return "Please log in as a student first. After that, I can read your own reservation records and show the latest status, schedule, room, and admin remarks."
        reservations = context.get('reservations', [])
        if not reservations:
            return "I did not find any reservation requests under your account yet.<br><br><strong>Next step:</strong> go to New Reservation, choose a facility and schedule, then submit the request for admin approval."
        first = reservations[0]
        remarks = f" Remarks: {first['remarks']}" if first.get('remarks') else ''
        return f"Here is your latest reservation:<br><br><strong>Request #{first['id']}</strong><br>Resource: {first['resource']}<br>Date: {first['date']}<br>Time: {first['time']}<br>Status: <strong>{first['status']}</strong>{remarks}<br><br>If it is pending, wait for admin review. If it is rejected, check the remarks and submit a corrected request."
    if intent == 'admin_insights':
        if context.get('not_allowed'):
            return context['not_allowed']
        return f"<strong>Admin activity summary</strong><br><br>Today: {context['today_total']} reservation(s)<br>Pending today: {context['today_pending']}<br>All pending approvals: {context['all_pending']}<br>Approved reservations: {context['all_approved']}<br>Active booking load: {context['active_booking_count']}<br><br>Most requested this month: <strong>{context['most_requested_facility_this_month']}</strong>.<br><br><strong>Suggested action:</strong> review pending approvals first, especially requests scheduled for today or peak hours."
    if intent == 'room_info':
        facilities = context.get('facilities', [])
        if not facilities:
            return "I could not find matching room records in the system."
        item = facilities[0]
        return f"<strong>{item.get('code')}</strong> is currently listed as <strong>{item.get('status')}</strong>.<br><br>Capacity: {item.get('capacity', 'N/A')}<br>Equipment: {item.get('equipment', 'N/A')}<br><br>For exact availability, ask with a date and time, for example: \"Is AVR available tomorrow 10 AM to 12 PM?\""
    if intent == 'requirements':
        return "<strong>Reservation requirements</strong><br><br>1. Active student account.<br>2. Facility or projector category.<br>3. Date, start time, and end time within 7:00 AM to 9:00 PM.<br>4. Available room or projector selected in the system.<br>5. Course code, course name, and professor.<br>6. Academic purpose or supporting details if the admin asks for clarification.<br><br>After submitting, your request becomes <strong>PENDING APPROVAL</strong> until an admin approves or rejects it."
    if intent == 'guide':
        return "<strong>How to reserve a facility</strong><br><br>1. Open <strong>New Reservation</strong>.<br>2. Choose the facility category: AVR, Computer Laboratory, Hospitality Management Laboratory, or Projector.<br>3. Select the date and time. The system checks conflicts using approved and pending requests.<br>4. Pick an available room or projector.<br>5. Enter course code, course name, and professor.<br>6. Submit the request and wait for admin review.<br><br>Status will show as <strong>PENDING APPROVAL</strong> until the admin approves or rejects it."
    if intent == 'manage':
        return "<strong>Cancel or modify a reservation</strong><br><br>If cancellation is available, open your reservation list and cancel the request there. To modify details like room, time, course, or professor, submit a corrected reservation after cancelling the old one if allowed.<br><br>If the request was already reviewed, check the admin remarks first so you know what to fix."
    suggestions = context.get('fallback_suggestions', CHATBOT_SUGGESTIONS['guest'])
    return "I am not sure about that yet, but I can help with reservation tasks.<br><br>Try asking:<br>" + '<br>'.join(f"- {s}" for s in suggestions)

def _call_cerebras_chat(message, history, context):
    api_key = os.environ.get('CEREBRAS_API_KEY')
    if not api_key:
        return None
    system_prompt = """You are the AI assistant of the PUP Reservation System.

Your role is to help students, faculty, and administrators with facility reservations.
You can explain reservation rules, guide users through booking steps, check availability, summarize booking status, and help admins review reservation activity.
Give helpful, detailed, but compact answers. Prefer short sections with labels such as Status, Details, and Next step.
If the user asks about availability, reservation status, schedules, or rooms, use only the provided SYSTEM_CONTEXT database snapshot before answering.
If the information is unavailable, say so clearly and suggest the next action.
For availability, mention the checked date, time, available resource codes, blocking statuses, and next booking step.
For requirements, list the exact booking fields and approval process.
For admins, summarize counts and suggest what to review next.
For greetings, thanks, and short acknowledgements such as "hi", "okay", or "got it", reply conversationally and briefly, then offer one relevant reservation action.
Do not invent reservation data."""
    messages = [{'role': 'system', 'content': system_prompt}]
    for item in (history or [])[-6:]:
        role = item.get('role')
        if role in ('user', 'assistant') and item.get('content'):
            messages.append({'role': role, 'content': _clean_chat_html(str(item['content']))[:600]})
    messages.append({
        'role': 'user',
        'content': f"SYSTEM_CONTEXT:\n{json.dumps(context, ensure_ascii=False, default=str)}\n\nUSER_MESSAGE:\n{message}"
    })
    payload = {
        'model': os.environ.get('CEREBRAS_MODEL', 'llama-3.3-70b'),
        'messages': messages,
        'temperature': 0.2,
        'max_tokens': 520
    }
    req = urllib.request.Request(
        'https://api.cerebras.ai/v1/chat/completions',
        data=json.dumps(payload).encode('utf-8'),
        headers={'Content-Type': 'application/json', 'Authorization': f"Bearer {api_key}"},
        method='POST'
    )
    try:
        with urllib.request.urlopen(req, timeout=12) as res:
            body = json.loads(res.read().decode('utf-8'))
        return body['choices'][0]['message']['content'].strip()
    except (urllib.error.URLError, urllib.error.HTTPError, KeyError, IndexError, json.JSONDecodeError, TimeoutError):
        return None

@ai_bp.route('/ai/chat', methods=['POST'])
def ai_chat():
    data = request.json or {}
    raw_message = (data.get('message') or '').strip()
    history = data.get('history') or []
    user = get_current_user()
    role = _chat_role(user)

    if not raw_message:
        return jsonify({
            'response': "Mabuhay! I can help with rooms, schedules, booking status, and reservation steps.",
            'suggestions': CHATBOT_SUGGESTIONS.get(role, CHATBOT_SUGGESTIONS['guest'])
        })

    intent = _detect_chat_intent(raw_message)
    context = _build_chat_context(user, raw_message, intent)
    ai_response = _call_cerebras_chat(raw_message, history, context) or _basic_chatbot_response(context)

    return jsonify({
        'response': ai_response,
        'intent': intent,
        'suggestions': CHATBOT_SUGGESTIONS.get(role, CHATBOT_SUGGESTIONS['guest'])
    })

    message = (data.get('message') or '').lower()
    
    response = "I am the PUP Parañaque AI Assistant. I can help you with campus resource status and reservation procedures. What would you like to know?"
    
    if 'how' in message and 'reserve' in message:
        response = "To reserve a facility:<br>1. Log in as a Student.<br>2. Navigate to <strong>New Reservation</strong>.<br>3. Fill in the category, date, operating times, and subject details.<br>4. Click <strong>Submit Booking Request</strong>. Admin will review it shortly."
    elif 'avr' in message or 'visual' in message:
        response = "The Audio-Visual Room (AVR-01) is located in the Main Building. It is equipped with airconditioning and audio controllers. You can reserve it for academic assemblies and lectures."
    elif 'computer' in message or 'lab' in message or 'laboratory' in message:
        response = "Our <strong>Computer Laboratory</strong> (COMP-LAB-01) features high-specification PCs, and our <strong>Hospitality Management Laboratory</strong> (HM-LAB-01) supports students for practical modules."
    elif 'projector' in message:
        response = "Classroom projectors (PJ-001 to PJ-010) are available in our portable fleet. Students can select available units interactively using the projector selection grid on the booking form."
    elif 'status' in message or 'pending' in message or 'approve' in message:
        response = "You can track and review the progress of all your submitted booking requests under the <strong>Availability & Status</strong> or <strong>History Log</strong> pages."
    elif 'mabuhay' in message or 'hello' in message or 'hi' in message:
        response = "Mabuhay! I am your PUP Parañaque AI Assistant. How can I help you today?"
        
    return jsonify({'response': response})
